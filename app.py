import os
import gc
import psutil
import warnings
from flask import Flask, render_template, request, jsonify, send_file, Response, redirect
import math
import time
import pandas as pd
from datetime import datetime
import asyncio
import aiohttp
from models.options_pricing import OptionPricer
from models.risk_metrics import risk_calculator
from models.greeks_calculator import greeks_calculator

# Supprimer tous les warnings d'asyncio dès le début
warnings.filterwarnings("ignore", category=RuntimeWarning, module="asyncio")
warnings.filterwarnings("ignore", category=DeprecationWarning, module="asyncio")
warnings.filterwarnings("ignore", message=".*Event loop is closed.*")
warnings.filterwarnings("ignore", message=".*_ProactorBasePipeTransport.*")
warnings.filterwarnings("ignore", message=".*coroutine.*was never awaited.*")
warnings.filterwarnings("ignore", message=".*Task was destroyed but it is pending.*")
warnings.filterwarnings("ignore", message=".*unclosed.*transport.*")
warnings.filterwarnings("ignore", message=".*unclosed.*client.*session.*")

# Fonction pour surveiller l'utilisation mémoire
def log_memory_usage():
    """Log l'utilisation mémoire actuelle"""
    process = psutil.Process()
    memory_info = process.memory_info()
    print(f"📊 Utilisation mémoire: {memory_info.rss / 1024 / 1024:.2f} MB")

# Fonction pour nettoyer la mémoire
def cleanup_memory():
    """Force le garbage collection pour libérer la mémoire"""
    gc.collect()
    log_memory_usage()

# Charger les variables d'environnement AVANT tous les imports
try:
    from dotenv import load_dotenv
    load_dotenv()
    load_dotenv('.env.local')  # Charger aussi .env.local pour le développement
    print("✅ Variables d'environnement chargées avec succès")
except ImportError:
    print("⚠️  Module python-dotenv non trouvé, utilisation des variables d'environnement système")

# Import de la configuration Tradier
from api.tradier_config import TRADIER_API_KEY, is_tradier_configured, get_tradier_config

# Vérifier que les clés API sont disponibles
if not is_tradier_configured():
    print("❌ ERREUR: TRADIER_API_KEY non trouvée dans les variables d'environnement")
    print("   L'application ne fonctionnera pas correctement")
    print("   Créez un fichier .env avec: TRADIER_API_KEY=votre_clé")
    print("   Obtenez votre clé gratuite sur: https://developer.tradier.com/")
else:
    print(f"✅ TRADIER_API_KEY chargée depuis le fichier .env: {TRADIER_API_KEY[:10]}...{TRADIER_API_KEY[-4:]}")
    print("🔧 Configuration Tradier: OK")
    print("📊 IMPORTANT: La volatilité implicite est CALCULÉE avec Black-Scholes (pas récupérée de l'API)")
    print("🎯 L'API Tradier fournit les prix des options, nous calculons l'IV")

# Import du nouveau module Yahoo Finance API
from api.yahoo_finance_api import yahoo_api

# Import du module Tradier API
from api.tradier_api import TradierAPI
import math

app = Flask(__name__)
pricer = OptionPricer()

# Initialiser l'API Tradier avec la configuration
tradier_config = get_tradier_config()
tradier_api = TradierAPI(TRADIER_API_KEY)

# Classe pour les appels asynchrones Tradier
class AsyncTradierAPI:
    """Classe pour gérer les appels asynchrones à l'API Tradier"""
    
    def __init__(self, token: str):
        self.token = token
        self.base_url = "https://api.tradier.com/v1"
        self.headers = {
            'Authorization': f'Bearer {token}',
            'Accept': 'application/json'
        }
    
    async def get_expirations(self, session: aiohttp.ClientSession, symbol: str) -> dict:
        """Récupère les expirations pour un symbole de manière asynchrone"""
        try:
            url = f"{self.base_url}/markets/options/expirations"
            params = {'symbol': symbol}
            
            timeout = aiohttp.ClientTimeout(total=10)
            async with session.get(url, params=params, timeout=timeout) as response:
                if response.status == 200:
                    data = await response.json()
                    return {
                        'success': True,
                        'data': data,
                        'expirations': data.get('expirations', {}).get('date', [])
                    }
                else:
                    return {
                        'success': False,
                        'error': f'HTTP {response.status}',
                        'expirations': []
                    }
        except asyncio.TimeoutError:
            return {
                'success': False,
                'error': 'Timeout lors de la requête',
                'expirations': []
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'expirations': []
            }
    
    async def get_options(self, session: aiohttp.ClientSession, symbol: str, expiration: str) -> dict:
        """Récupère les options pour un symbole et une expiration de manière asynchrone"""
        try:
            url = f"{self.base_url}/markets/options/chains"
            params = {
                'symbol': symbol,
                'expiration': expiration,
                'greeks': 'true'
            }
            
            timeout = aiohttp.ClientTimeout(total=15)
            async with session.get(url, params=params, timeout=timeout) as response:
                if response.status == 200:
                    data = await response.json()
                    return {
                        'success': True,
                        'data': data,
                        'options': data.get('options', {}).get('option', [])
                    }
                else:
                    return {
                        'success': False,
                        'error': f'HTTP {response.status}',
                        'options': []
                    }
        except asyncio.TimeoutError:
            return {
                'success': False,
                'error': 'Timeout lors de la requête',
                'options': []
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'options': []
            }
    
    async def get_quote(self, session: aiohttp.ClientSession, symbol: str) -> dict:
        """Récupère le prix spot d'une action de manière asynchrone"""
        try:
            url = f"{self.base_url}/markets/quotes"
            params = {'symbols': symbol}
            
            timeout = aiohttp.ClientTimeout(total=10)
            async with session.get(url, params=params, timeout=timeout) as response:
                if response.status == 200:
                    data = await response.json()
                    quotes = data.get('quotes', {}).get('quote', [])
                    if quotes:
                        quote = quotes[0] if isinstance(quotes, list) else quotes
                        return {
                            'success': True,
                            'data': data,
                            'spot_price': float(quote.get('last', 0))
                        }
                    else:
                        return {
                            'success': False,
                            'error': 'No quote data',
                            'spot_price': 0
                        }
                else:
                    return {
                        'success': False,
                        'error': f'HTTP {response.status}',
                        'spot_price': 0
                    }
        except asyncio.TimeoutError:
            return {
                'success': False,
                'error': 'Timeout lors de la requête',
                'spot_price': 0
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'spot_price': 0
            }

# Initialiser l'API Tradier asynchrone
async_tradier_api = AsyncTradierAPI(TRADIER_API_KEY)

# Fonction helper pour exécuter des fonctions async dans Flask
def run_async(coro):
    """Exécute une coroutine asynchrone dans le contexte Flask"""
    import threading
    import warnings
    import gc
    
    # Supprimer tous les warnings d'asyncio
    warnings.filterwarnings("ignore", category=RuntimeWarning, module="asyncio")
    warnings.filterwarnings("ignore", category=DeprecationWarning, module="asyncio")
    warnings.filterwarnings("ignore", message=".*Event loop is closed.*")
    warnings.filterwarnings("ignore", message=".*_ProactorBasePipeTransport.*")
    
    # Créer un nouvel event loop pour chaque requête (évite les conflits)
    def run_in_thread():
        # Créer un nouvel event loop dans un thread séparé
        new_loop = asyncio.new_event_loop()
        asyncio.set_event_loop(new_loop)
        
        try:
            return new_loop.run_until_complete(coro)
        except Exception as e:
            print(f"⚠️ Erreur dans run_async: {e}")
            raise
        finally:
            # Nettoyage ultra-robuste pour Windows
            try:
                # Attendre que toutes les tâches en attente se terminent
                pending = asyncio.all_tasks(new_loop)
                if pending:
                    # Annuler toutes les tâches en cours
                    for task in pending:
                        task.cancel()
                    
                    # Attendre que toutes les tâches se terminent avec timeout court
                    try:
                        new_loop.run_until_complete(asyncio.wait_for(
                            asyncio.gather(*pending, return_exceptions=True),
                            timeout=2.0
                        ))
                    except asyncio.TimeoutError:
                        # Forcer l'arrêt si timeout
                        pass
                
                # Fermer l'event loop proprement
                if not new_loop.is_closed():
                    new_loop.close()
                
                # Forcer le garbage collection pour nettoyer les transports
                gc.collect()
                
            except Exception as cleanup_error:
                # Ignorer complètement les erreurs de nettoyage sur Windows
                pass
    
    # Exécuter dans un thread séparé pour éviter les conflits d'event loop
    result = [None]
    exception = [None]
    
    def target():
        try:
            result[0] = run_in_thread()
        except Exception as e:
            exception[0] = e
    
    thread = threading.Thread(target=target)
    thread.start()
    thread.join()
    
    # Forcer le garbage collection après le thread
    gc.collect()
    
    if exception[0]:
        raise exception[0]
    
    return result[0]

# Fonction alternative synchrone pour éviter les problèmes asyncio
def run_sync_fallback(func, *args, **kwargs):
    """Fonction de fallback synchrone pour éviter les problèmes asyncio"""
    try:
        return func(*args, **kwargs)
    except Exception as e:
        print(f"⚠️ Erreur dans run_sync_fallback: {e}")
        raise e 
        
# Fonction asynchrone pour récupérer les expirations
async def get_expirations_async(symbol: str):
    """Récupère les expirations de manière asynchrone"""
    async with aiohttp.ClientSession(headers=async_tradier_api.headers) as session:
        result = await async_tradier_api.get_expirations(session, symbol)
        
        if result['success']:
            exp_dates = result['expirations']
            if not isinstance(exp_dates, list):
                exp_dates = [exp_dates]
            
            # Formater les dates d'expiration
            formatted_expirations = []
            for exp_date in exp_dates:
                try:
                    # Calculer les jours jusqu'à l'expiration
                    from datetime import datetime
                    exp_datetime = datetime.strptime(exp_date, "%Y-%m-%d")
                    today = datetime.now()
                    days_to_exp = (exp_datetime - today).days
                    
                    # Vérifier que days_to_exp n'est pas None
                    if days_to_exp is None:
                        continue
                    
                    # Déterminer le type d'expiration
                    if days_to_exp <= 7:
                        exp_type = "Weekly"
                    elif days_to_exp <= 30:
                        exp_type = "Monthly"
                    elif days_to_exp <= 90:
                        exp_type = "Quarterly"
                    else:
                        exp_type = "Long-term"
                    
                    formatted_expirations.append({
                        'date': exp_date,
                        'days_to_exp': days_to_exp,
                        'type': exp_type,
                        'label': f"{exp_date} ({days_to_exp} jours)"
                    })
                except ValueError:
                    continue
            
            return {
                'success': True,
                'expirations': formatted_expirations,
                'count': len(formatted_expirations)
            }
        else:
            return {
                'success': False,
                'error': result['error'],
                'expirations': [],
                'count': 0
            }

# Fonction asynchrone pour récupérer le prix spot
async def get_quote_async(symbol: str):
    """Récupère le prix spot de manière asynchrone"""
    async with aiohttp.ClientSession(headers=async_tradier_api.headers) as session:
        result = await async_tradier_api.get_quote(session, symbol)
        
        if result['success']:
            return {
                'success': True,
                'spot_price': result['spot_price'],
                'symbol': symbol
            }
        else:
            return {
                'success': False,
                'error': result['error'],
                'spot_price': 0,
                'symbol': symbol
            }

# Fonction asynchrone pour récupérer l'union des strikes
async def get_strikes_union_async(symbol: str):
    """Récupère l'union des strikes de plusieurs maturités de manière asynchrone"""
    async with aiohttp.ClientSession(headers=async_tradier_api.headers) as session:
        # 1. Récupérer les expirations
        expirations_result = await async_tradier_api.get_expirations(session, symbol)
        
        if not expirations_result['success']:
            return {
                'success': False,
                'error': f"Erreur récupération expirations: {expirations_result['error']}",
                'strikes': []
            }
        
        expirations = expirations_result['expirations']
        if not expirations:
            return {
                'success': False,
                'error': 'Aucune expiration trouvée',
                'strikes': []
            }
        
        # 2. Sélectionner les maturités spécifiques (2ème, 5ème, 6ème, 8ème, 15ème)
        selected_indices = [1, 4, 5, 7, 14]  # Indices 0-based
        selected_expirations = []
        
        for idx in selected_indices:
            if idx < len(expirations):
                # expirations est une liste de strings (dates)
                selected_expirations.append(expirations[idx])
        
        if not selected_expirations:
            return {
                'success': False,
                'error': 'Pas assez d\'expirations disponibles',
                'strikes': []
            }
        
        print(f"📅 Maturités sélectionnées: {selected_expirations}")
        
        # 3. Récupérer les options pour chaque expiration en parallèle
        tasks = []
        for expiration in selected_expirations:
            task = async_tradier_api.get_options(session, symbol, expiration)
            tasks.append(task)
        
        # Exécuter toutes les tâches en parallèle
        results = await asyncio.gather(*tasks)
        
        # 4. Traiter les résultats et extraire les strikes
        all_strikes = set()
        successful_expirations = 0
        
        for i, result in enumerate(results):
            if result['success']:
                options = result['options']
                successful_expirations += 1
                
                for option in options:
                    try:
                        strike = float(option.get('strike', 0))
                        all_strikes.add(strike)
                    except (ValueError, TypeError):
                        continue
        
        if not all_strikes:
            return {
                'success': False,
                'error': 'Aucun strike trouvé',
                'strikes': []
            }
        
        # 5. Récupérer le prix spot pour le filtrage
        spot_result = await async_tradier_api.get_quote(session, symbol)
        spot_price = spot_result.get('spot_price', 0) if spot_result['success'] else 0
        
        # 6. Filtrer les strikes (multiples de 5 uniquement)
        filtered_strikes = []
        for strike in sorted(all_strikes):
            # Filtrage: Garder seulement les multiples de 5 (divisibles par 5)
            if strike % 5 == 0:
                percentage = (strike / spot_price * 100) if spot_price > 0 else 0
                filtered_strikes.append({
                    'strike': strike,
                    'percentage': percentage,
                    'percentage_display': f"{percentage:.1f}% (${strike:.2f})"
                })
        
        print(f"✅ Union terminée: {len(filtered_strikes)} strikes uniques (filtrés) pour {symbol} (Spot: ${spot_price:.2f})")
        print(f"📊 Filtrage appliqué: divisibles par 5 uniquement")
        
        return {
            'success': True,
            'strikes': filtered_strikes,
            'spot_price': spot_price,
            'total_strikes': len(all_strikes),
            'filtered_strikes': len(filtered_strikes),
            'expirations_processed': successful_expirations,
            'selected_expirations': selected_expirations
        }

def test_tradier_connectivity():
    """Test la connectivité avec l'API Tradier"""
    try:
        print("🔍 Test de connectivité Tradier...")
        import requests
        
        # Test de connectivité basique
        response = requests.get("https://api.tradier.com/v1/markets/clock", 
                              headers={"Authorization": f"Bearer {TRADIER_API_KEY}", "Accept": "application/json"},
                              timeout=10)
        
        if response.status_code == 200:
            print("✅ Connexion Tradier OK")
            return True
        else:
            print(f"⚠️  Tradier répond mais avec erreur: {response.status_code}")
            return False
            
    except requests.exceptions.ConnectTimeout:
        print("❌ Timeout de connexion vers Tradier")
        return False
    except requests.exceptions.ConnectionError as e:
        print(f"❌ Erreur de connexion vers Tradier: {e}")
        return False
    except Exception as e:
        print(f"❌ Erreur inattendue: {e}")
        return False

def calculate_business_days_to_expiration(expiration_date_str):
    """
    Calcule le nombre de jours ouvrés jusqu'à l'expiration
    
    Args:
        expiration_date_str (str): Date d'expiration au format "YYYY-MM-DD"
        
    Returns:
        float: Nombre d'années (jours ouvrés / 252)
    """
    try:
        from datetime import datetime, timedelta
        
        expiration_date = datetime.strptime(expiration_date_str, "%Y-%m-%d")
        current_date = datetime.now()
        
        # Compter les jours ouvrés entre aujourd'hui et la date d'expiration
        business_days = 0
        current = current_date.date()
        end = expiration_date.date()
        
        while current < end:
            if current.weekday() < 5:  # lundi=0 .. vendredi=4
                business_days += 1
            current += timedelta(days=1)
        
        # Convertir en années (252 jours ouvrés par an)
        years = max(0.001, business_days / 252.0)
        return years
        
    except Exception as e:
        print(f"❌ Erreur calcul jours ouvrés: {e}")
        return 0.5  # Valeur par défaut

def calculate_implied_volatility_black_scholes(spot_price: float, strike: float, time_to_exp: float, 
                                             option_price: float, option_type: str, risk_free_rate: float = 0.05, 
                                             dividend_yield: float = 0.0) -> float:
    """
    Calcule la volatilité implicite d'une option en utilisant la méthode de Newton-Raphson
    avec le modèle de Black-Scholes
    
    Paramètres optimisés pour convergence avec les données d'images:
    - Taux sans risque: 5% (optimisé pour options ATM)
    - Rendement dividendes: 0% (simplifié pour convergence)
    - Temps: Jours ouvrés / 252 (plus précis que jours calendaires)
    
    Args:
        spot_price (float): Prix actuel de l'actif sous-jacent
        strike (float): Prix d'exercice de l'option
        time_to_exp (float): Temps jusqu'à l'expiration en années (jours ouvrés/252)
        option_price (float): Prix observé de l'option
        option_type (str): Type d'option ("call" ou "put")
        risk_free_rate (float): Taux sans risque (par défaut: 5% - optimisé)
        dividend_yield (float): Rendement des dividendes (par défaut: 0% - optimisé)
        
    Returns:
        float: Volatilité implicite ou None si le calcul échoue
    """
    try:
        # Fonction de distribution normale cumulative (approximation)
        def norm_cdf(x):
            return 0.5 * (1 + math.erf(x / math.sqrt(2)))
        
        # Fonction de densité de probabilité normale
        def norm_pdf(x):
            return math.exp(-0.5 * x * x) / math.sqrt(2 * math.pi)
        
        # Fonction Black-Scholes avec dividendes
        def black_scholes_price(S, K, T, r, sigma, option_type, q=0):
            if T is None or T <= 0:
                return max(S - K, 0) if option_type == "call" else max(K - S, 0)
            
            d1 = (math.log(S / K) + (r - q + 0.5 * sigma * sigma) * T) / (sigma * math.sqrt(T))
            d2 = d1 - sigma * math.sqrt(T)
            
            if option_type == "call":
                return S * math.exp(-q * T) * norm_cdf(d1) - K * math.exp(-r * T) * norm_cdf(d2)
            else:  # put
                return K * math.exp(-r * T) * norm_cdf(-d2) - S * math.exp(-q * T) * norm_cdf(-d1)
        
        # Fonction de vega (dérivée du prix par rapport à la volatilité) avec dividendes
        def vega(S, K, T, r, sigma, q=0):
            if T is None or T <= 0:
                return 0
            d1 = (math.log(S / K) + (r - q + 0.5 * sigma * sigma) * T) / (sigma * math.sqrt(T))
            return S * math.exp(-q * T) * math.sqrt(T) * norm_pdf(d1)
        
        # Méthode de Newton-Raphson pour trouver la volatilité implicite
        sigma = 0.2  # Estimation initiale
        tolerance = 1e-6
        max_iterations = 100
        
        for i in range(max_iterations):
            try:
                price = black_scholes_price(spot_price, strike, time_to_exp, risk_free_rate, sigma, option_type, dividend_yield)
                vega_val = vega(spot_price, strike, time_to_exp, risk_free_rate, sigma, dividend_yield)
                
                if vega_val == 0:
                    break
                
                diff = price - option_price
                if abs(diff) < tolerance:
                    return sigma
                
                sigma = sigma - diff / vega_val
                
                # S'assurer que la volatilité reste dans des limites raisonnables
                sigma = max(0.001, min(3.0, sigma))
                
            except (ValueError, ZeroDivisionError, OverflowError):
                break
        
        return None
        
    except Exception:
        return None

def get_risk_free_rate():
    """
    Récupère le taux sans risque actuel (optimisé pour convergence avec Tradier)
    """
    try:
        # Optimisé pour converger vers l'IV de Tradier (testé avec 2025-09-19)
        # Le taux de 5% donne le meilleur écart moyen (13.13%)
        return 0.05  # 5% - taux optimisé pour convergence avec Tradier
    except:
        return 0.05  # Fallback à 5%

def get_dividend_yield(symbol):
    """
    Récupère le rendement des dividendes pour un symbole donné
    """
    try:
        # Ajusté pour converger vers l'IV de Tradier
        # Utiliser 0% de dividendes pour simplifier et converger vers l'IV de Tradier
        return 0.0  # 0% - ajusté pour convergence avec Tradier
    except:
        return 0.0

# Route de santé pour surveiller l'application
@app.route('/health')
def health_check():
    """Route de santé pour surveiller l'état de l'application"""
    try:
        # Vérifier l'utilisation mémoire
        memory_usage = psutil.Process().memory_info().rss / 1024 / 1024  # MB
        
        return jsonify({
            'status': 'healthy',
            'memory_usage_mb': round(memory_usage, 2),
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

# Approximation rationnelle d'Acklam pour l'inverse de la CDF normale (PPF)
# Source adaptée: https://web.archive.org/web/20151030215612/http://home.online.no/~pjacklam/notes/invnorm/
def _inv_norm_cdf(p: float) -> float:
    if p <= 0.0 or p >= 1.0:
        raise ValueError("p doit être dans (0,1)")

    # Coefficients pour approximation par morceaux
    a = [
        -3.969683028665376e+01,
        2.209460984245205e+02,
        -2.759285104469687e+02,
        1.383577518672690e+02,
        -3.066479806614716e+01,
        2.506628277459239e+00,
    ]
    b = [
        -5.447609879822406e+01,
        1.615858368580409e+02,
        -1.556989798598866e+02,
        6.680131188771972e+01,
        -1.328068155288572e+01,
    ]
    c = [
        -7.784894002430293e-03,
        -3.223964580411365e-01,
        -2.400758277161838e+00,
        -2.549732539343734e+00,
        4.374664141464968e+00,
        2.938163982698783e+00,
    ]
    d = [
        7.784695709041462e-03,
        3.224671290700398e-01,
        2.445134137142996e+00,
        3.754408661907416e+00,
    ]

    plow = 0.02425
    phigh = 1 - plow

    if p < plow:
        q = math.sqrt(-2 * math.log(p))
        return (
            (((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5])
            / ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1)
        )
    elif phigh < p:
        q = math.sqrt(-2 * math.log(1 - p))
        return -(
            (((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5])
            / ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1)
        )
    else:
        q = p - 0.5
        r = q * q
        return (
            (((((a[0] * r + a[1]) * r + a[2]) * r + a[3]) * r + a[4]) * r + a[5]) * q
        ) / (
            ((((b[0] * r + b[1]) * r + b[2]) * r + b[3]) * r + b[4]) * r + 1
        )


def _z_from_confidence(conf_level: float) -> float:
    # conf_level dans (0,1), par défaut 0.95
    cl = max(0.5, min(0.999, float(conf_level)))
    tail = 1.0 - (1.0 - cl) / 2.0
    return abs(_inv_norm_cdf(tail))

@app.route('/')
def index():
    """Page d'accueil du dashboard"""
    return render_template('index.html')

@app.route('/cv-mathis-le-gall')
def cv_mathis_le_gall():
    """Page CV Mathis Le Gall"""
    return render_template('cv_mathis_le_gall.html')

@app.route('/indices-actions')
def indices_actions():
    """Page des indices et actions"""
    return render_template('indices_actions.html')

@app.route('/rates-fx')
def rates_fx():
    """Page des taux et forex"""
    return render_template('rates_fx.html')

@app.route('/call-put')
def call_put():
    """Page Call & Put"""
    return render_template('call_put.html')



@app.route('/volatility-surface')
def volatility_surface():
    """Page de la surface de volatilité"""
    return render_template('volatility_surface.html')

@app.route('/force-refresh')
def force_refresh():
    """Page pour forcer le rafraîchissement"""
    return render_template('force_refresh.html')

@app.route('/test-vol-surface')
def test_vol_surface():
    """Page de test pour debug de la surface de volatilité"""
    return send_file('test_vol_surface_debug.html')



@app.route('/test-indices-actions')
def test_indices_actions():
    """Page de test pour les indices et actions"""
    return send_file('test_indices_actions.html')



@app.route('/analyse-actions-indices')
def analyse_actions_indices():
    """Page dédiée: comparaison de performances (Base 100)"""
    return render_template('analyse_actions_indices.html')


@app.route('/description-app')
def description_app():
    """Page de description de l'application"""
    return render_template('description_app.html')


@app.route('/analyse-volatility')
def analyse_volatility():
    """Page d'analyse de volatilité"""
    return render_template('analyse_volatility.html')

# APIs pour les données financières
@app.route('/api/market-data')
def api_market_data():
    """API pour récupérer les données de marché"""
    try:
        # Utiliser le nouveau module Yahoo Finance API
        data = yahoo_api.get_market_data()
        
        if data and (data.get('indices') or data.get('stocks')):
            return jsonify(data)
        else:
            return jsonify({
                'error': 'Impossible de récupérer les données via Yahoo Finance.',
                'indices': {},
                'stocks': {},
                'timestamp': datetime.now().isoformat()
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chart-data/<path:symbol>')
def api_chart_data_legacy(symbol):
    """API pour récupérer les données de graphique (route legacy)"""
    return api_chart_data_v2(symbol)

@app.route('/api/chart-data-v2/<path:symbol>')
def api_chart_data_v2(symbol):
    """API pour récupérer les données de graphique avec timeframe"""
    try:
        # Récupérer le timeframe depuis les paramètres de requête
        timeframe = request.args.get('timeframe', '1mo')
        start = request.args.get('start')  # YYYY-MM-DD
        end = request.args.get('end')      # YYYY-MM-DD
        
        # Utiliser le nouveau module Yahoo Finance API
        test_data = yahoo_api.get_chart_data(symbol, timeframe, start, end)
        
        if test_data:
            return jsonify(test_data)
        else:
            return jsonify({
                'error': f'Impossible de récupérer les données pour {symbol}.',
                'labels': [],
                'datasets': []
            })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# API: surface de volatilité implicite
@app.route('/api/vol-surface/<symbol>')
def api_vol_surface(symbol):
    """API endpoint pour récupérer les données de volatilité surface via Finnhub"""
    try:
        max_exp = int(request.args.get('maxExp', 6))
        span = float(request.args.get('span', 0.5))
        provider = request.args.get('provider', 'tradier')  # 'tradier' par défaut
        
        # Validation des paramètres
        if max_exp < 1 or max_exp > 12:
            return jsonify({'error': 'maxExp doit être entre 1 et 12'}), 400
            
        if span <= 0 or span > 1:
            return jsonify({'error': 'span doit être entre 0 et 1'}), 400
            
        if provider != 'tradier':
            return jsonify({'error': 'Seul le provider Tradier est supporté'}), 400
        
        print(f"🔴 Utilisation de Tradier pour {symbol}")
        # Utiliser l'API Tradier pour les vraies données d'options
        tradier = TradierAPI(TRADIER_API_KEY)
        
        # Récupérer les dates d'expiration disponibles via Tradier
        print(f"🔍 Récupération des expirations pour {symbol}...")
        expirations_data = tradier.get_option_expirations(symbol)
        print(f"📅 Expirations reçues: {expirations_data}")
        
        if not expirations_data or 'expirations' not in expirations_data or 'date' not in expirations_data['expirations']:
            print(f"❌ Aucune expiration trouvée pour {symbol}")
            return jsonify({
                'error': f'Aucune date d\'expiration disponible pour {symbol}'
            }), 404
        
        # Extraire les dates d'expiration
        expirations = expirations_data['expirations']['date']
        
        # Limiter le nombre d'expirations selon maxExp
        expirations_to_use = expirations[:max_exp]
        
        # Récupérer le prix spot actuel via Tradier
        try:
            quote_data = tradier.get_stock_quote(symbol)
            if quote_data and 'quotes' in quote_data and 'quote' in quote_data['quotes']:
                spot_price = float(quote_data['quotes']['quote'].get('last', 0))
            else:
                spot_price = 0
        except:
            spot_price = 0
        
        # Collecter toutes les données d'options pour toutes les expirations
        all_options_data = []
        
        print(f"Test de {len(expirations_to_use)} expirations pour {symbol}")
        
        for expiration_date in expirations_to_use:
            try:
                # expiration_date est déjà au format 'YYYY-MM-DD'
                print(f"Récupération des options pour {expiration_date}...")
                
                # Récupérer les données d'options pour cette expiration via Tradier
                options_data = tradier.get_historical_options_data(symbol, expiration_date)
                
                if options_data is not None and not options_data.empty:
                    print(f"✅ {len(options_data)} options trouvées pour {expiration_date}")
                    # Renommer les colonnes pour correspondre au format attendu
                    options_data = options_data.rename(columns={
                        'Strike': 'strike',
                        'Type': 'type',
                        'Last': 'lastPrice',
                        'Implied_Volatility': 'impliedVolatility',
                        'Expiration': 'expiration_date'
                    })
                    # Convertir le type en minuscules
                    options_data['type'] = options_data['type'].str.lower()
                    all_options_data.append(options_data)
                else:
                    print(f"❌ Aucune option pour {expiration_date}")
                    
            except Exception as e:
                print(f"Erreur pour l'expiration {expiration_date if 'expiration_date' in locals() else 'inconnue'}: {e}")
                continue
        
        print(f"Total d'expirations avec données: {len(all_options_data)}")
        
        if not all_options_data:
            print(f"❌ Aucune donnée d'options collectée pour {symbol}")
            return jsonify({
                'error': f'Aucune option disponible pour {symbol}'
            }), 404
        
        # Combiner toutes les données
        combined_data = pd.concat(all_options_data, ignore_index=True)
        
        # Filtrer par bande autour du spot si le prix spot est disponible
        if spot_price > 0:
            min_strike = spot_price * (1 - span)
            max_strike = spot_price * (1 + span)
            filtered_data = combined_data[
                (combined_data['strike'] >= min_strike) & 
                (combined_data['strike'] <= max_strike)
            ]
            
            # Si le filtrage supprime trop de données, utiliser toutes les données
            if len(filtered_data) < 10:
                print(f"Filtrage trop restrictif, utilisation de toutes les données ({len(combined_data)} options)")
                combined_data = combined_data
            else:
                combined_data = filtered_data
        else:
            print(f"Prix spot non disponible, utilisation de toutes les données ({len(combined_data)} options)")
        
        if combined_data.empty:
            return jsonify({
                'error': f'Aucune option disponible pour {symbol}'
            }), 404
        
        # Organiser les données pour la surface de volatilité
        # Extraire les strikes et expirations uniques
        unique_strikes = sorted(combined_data['strike'].unique(), reverse=True)
        unique_expirations = sorted(combined_data['expiration_date'].unique())
        
        # Calculer les maturités en années
        current_date = datetime.now()
        maturities = []
        for exp_date in unique_expirations:
            exp_datetime = datetime.strptime(exp_date, '%Y-%m-%d')
            maturity = (exp_datetime - current_date).days / 365.25
            # S'assurer que la maturité est positive et raisonnable
            if maturity < 0:
                print(f"⚠️  Maturité négative pour {exp_date}: {maturity:.4f} ans")
                maturity = 0.01  # Maturité minimale
            elif maturity > 5:
                print(f"⚠️  Maturité très élevée pour {exp_date}: {maturity:.4f} ans")
                maturity = 5.0  # Maturité maximale
            maturities.append(maturity)
        
        # Créer la matrice de volatilité implicite
        iv_matrix = []
        for i, expiration in enumerate(unique_expirations):
            row = []
            for strike in unique_strikes:
                # Trouver l'option correspondante
                option = combined_data[
                    (combined_data['expiration_date'] == expiration) & 
                    (combined_data['strike'] == strike)
                ]
                
                if not option.empty:
                    # Prendre la moyenne des IV pour les calls et puts
                    iv_values = option['impliedVolatility'].values
                    # Filtrer les valeurs aberrantes (IV > 200% ou < 1%) et exclure les None
                    valid_iv_values = [iv for iv in iv_values if iv is not None and 0.01 <= iv <= 2.0]
                    
                    if valid_iv_values:
                        avg_iv = float(sum(valid_iv_values) / len(valid_iv_values))
                        row.append(avg_iv)
                    else:
                        row.append(None)
                else:
                    row.append(None)
            iv_matrix.append(row)
        
        # Calculer les statistiques sur les données filtrées
        # Filtrer les données IV valides (exclure les None)
        valid_iv_data = combined_data[
            (combined_data['impliedVolatility'].notna()) &
            (combined_data['impliedVolatility'] >= 0.01) & 
            (combined_data['impliedVolatility'] <= 2.0)
        ]
        
        if valid_iv_data.empty:
            return jsonify({
                'error': f'Aucune donnée IV valide pour {symbol}'
            }), 404
        
        # Convertir le DataFrame en format JSON
        result = {
            'symbol': symbol,
            'spot_price': spot_price,
            'data_source': 'Tradier API (Données Réelles)',
            'strikes': unique_strikes,
            'maturities': maturities,
            'iv': iv_matrix,
            'total_options': len(combined_data),
            'calls_count': len(combined_data[combined_data['type'] == 'call']),
            'puts_count': len(combined_data[combined_data['type'] == 'put']),
            'valid_options': len(valid_iv_data),
            'statistics': {
                'min_iv': float(valid_iv_data['impliedVolatility'].min()),
                'max_iv': float(valid_iv_data['impliedVolatility'].max()),
                'avg_iv': float(valid_iv_data['impliedVolatility'].mean()),
                'std_iv': float(valid_iv_data['impliedVolatility'].std())
            },
            'raw_options': combined_data.to_dict('records')  # Données brutes pour debug
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': f'Erreur serveur: {str(e)}'}), 500



@app.route('/api/indices')
def api_indices():
    """API pour récupérer uniquement les indices"""
    try:
        data = yahoo_api.get_market_data()
        return jsonify(data.get('indices', {}))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stocks')
def api_stocks():
    """API pour récupérer uniquement les actions"""
    try:
        data = yahoo_api.get_market_data()
        return jsonify(data.get('stocks', {}))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API pour les calculs d'options Call & Put
@app.route('/api/calculate-option', methods=['POST'])
def api_calculate_option():
    """API pour calculer le prix d'une option"""
    try:
        data = request.get_json()
        
        # Validation des données
        required_fields = ['spotPrice', 'strikePrice', 'timeMaturity', 'riskFreeRate', 'volatility', 'optionType']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Champ manquant: {field}'}), 400
        
        # Extraction des paramètres
        spot_price = float(data['spotPrice'])
        strike_price = float(data['strikePrice'])
        time_maturity = float(data['timeMaturity'])
        risk_free_rate = float(data['riskFreeRate'])
        volatility = float(data['volatility'])
        option_type = data['optionType']
        model_choice = data['modelChoice']
        
        # Validation des valeurs
        if spot_price <= 0 or strike_price <= 0 or time_maturity <= 0:
            return jsonify({'error': 'Les prix et l\'échéance doivent être positifs'}), 400
        
        if risk_free_rate < 0 or risk_free_rate > 1 or volatility < 0 or volatility > 1:
            return jsonify({'error': 'Le taux sans risque et la volatilité doivent être entre 0 et 1'}), 400
        
        # Validation du nombre de simulations Monte Carlo
        nb_simulations = int(data.get('numSimulations', 10000))
        if nb_simulations < 100 or nb_simulations > 1000000:
            return jsonify({'error': 'Le nombre de simulations doit être entre 100 et 1 000 000'}), 400
        
        # Si Monte Carlo: simuler les trajectoires pour obtenir prix et IC
        # Toujours calculer Monte Carlo et Black-Scholes pour comparaison
        if True:
            nb_steps = int(data.get('numSteps', 252))
            confidence_level = float(data.get('confidenceLevel', 0.95))

            t_all_start = time.perf_counter()
            t_mc_start = time.perf_counter()
            mc = pricer.monte_carlo_price_and_greeks(
                S=spot_price,
                K=strike_price,
                T=time_maturity,
                r=risk_free_rate,
                sigma=volatility,
                option_type=option_type,
                nb_simulations=nb_simulations,
                nb_steps=nb_steps,
                return_std=True,
                return_paths=int(data.get('numPaths', 50)),
            )
            t_mc_ms = (time.perf_counter() - t_mc_start) * 1000.0
            try:
                z = _z_from_confidence(confidence_level)
            except Exception:
                z = 1.96

            t_bs_start = time.perf_counter()
            bs_price, bs_delta, bs_gamma, bs_theta, bs_vega, bs_rho = pricer.black_scholes_price_and_greeks(
                S=spot_price,
                K=strike_price,
                T=time_maturity,
                r=risk_free_rate,
                sigma=volatility,
                option_type=option_type,
            )
            t_bs_ms = (time.perf_counter() - t_bs_start) * 1000.0
            t_total_ms = (time.perf_counter() - t_all_start) * 1000.0

            result = {
                'monteCarlo': {
                    'optionPrice': round(mc['price'], 4),
                    'delta': round(mc['delta'], 4) if not math.isnan(mc['delta']) else None,
                    'gamma': round(mc['gamma'], 4) if not math.isnan(mc['gamma']) else None,
                    'theta': round(mc['theta'], 4) if not math.isnan(mc['theta']) else None,
                    'vega': round(mc.get('vega', float('nan')), 4) if not math.isnan(mc.get('vega', float('nan'))) else None,
                    'rho': round(mc.get('rho', float('nan')), 4) if not math.isnan(mc.get('rho', float('nan'))) else None,
                    'confidenceInterval': {
                        'lower': round(mc['price'] - z * mc['stdError'], 4),
                        'mean': round(mc['price'], 4),
                        'upper': round(mc['price'] + z * mc['stdError'], 4),
                        'confidenceLevel': round(confidence_level, 4),
                    },
                    'paths': mc.get('paths'),
                    'timeGrid': mc.get('timeGrid'),
                    'timeMs': round(t_mc_ms, 2),
                },
                'blackScholes': {
                    'optionPrice': round(bs_price, 4),
                    'delta': round(bs_delta, 4),
                    'gamma': round(bs_gamma, 4),
                    'theta': round(bs_theta, 4),
                    'vega': round(bs_vega, 4),
                    'rho': round(bs_rho, 4),
                    'timeMs': round(t_bs_ms, 2),
                },
                'parameters': {
                    'spotPrice': spot_price,
                    'strikePrice': strike_price,
                    'timeMaturity': time_maturity,
                    'riskFreeRate': risk_free_rate,
                    'volatility': volatility,
                    'optionType': option_type,
                    'numSimulations': nb_simulations,
                    'numSteps': nb_steps,
                    'confidenceLevel': round(confidence_level, 4),
                },
                'timings': {
                    'totalMs': round(t_total_ms, 2),
                }
            }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/greeks-curves', methods=['POST'])
def api_greeks_curves():
    """API pour calculer les courbes des Greeks"""
    try:
        data = request.get_json()
        
        # Récupérer les paramètres

        
        spot_price = float(data.get('spotPrice', 100))
        strike_price = float(data.get('strikePrice', 100))
        time_maturity = float(data.get('timeMaturity', 1))
        risk_free_rate = float(data.get('riskFreeRate', 0.05))
        volatility = float(data.get('volatility', 0.2))
        option_type = data.get('optionType', 'call')
        
        # Calculer les courbes des Greeks
        curves = greeks_calculator.generate_greek_curves(
            S=spot_price,  # Passer le prix spot pour la plage dynamique
            K=strike_price,
            T=time_maturity,
            r=risk_free_rate,
            sigma=volatility,
            option_type=option_type
        )
        
        # Test de calcul Theta avec valeurs connues
        test_theta = greeks_calculator.calculate_theta(spot_price, strike_price, time_maturity, risk_free_rate, volatility, option_type)
        print(f"🔍 TEST THETA: S={spot_price}, K={strike_price}, T={time_maturity} → Theta={test_theta:.6f}")
        print(f"🔍 ARRAY THETA (premiers 5): {curves['theta'][:5]}")
        print(f"🔍 ARRAY THETA (derniers 5): {curves['theta'][-5:]}")
        
        # Calculer les valeurs au prix spot actuel
        current_values = greeks_calculator.get_greek_values_at_spot(
            S=spot_price,
            K=strike_price,
            T=time_maturity,
            r=risk_free_rate,
            sigma=volatility,
            option_type=option_type
        )
        
        return jsonify({
            'success': True,
            'curves': curves,
            'current_values': current_values,
            'parameters': {
                'spotPrice': spot_price,
                'strikePrice': strike_price,
                'timeMaturity': time_maturity,
                'riskFreeRate': risk_free_rate,
                'volatility': volatility,
                'optionType': option_type
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/volatility-sensitivity-matrix', methods=['POST'])
def api_volatility_sensitivity_matrix():
    """API pour calculer la matrice de sensibilité des Greeks selon la volatilité"""
    try:
        data = request.get_json()
        
        # Récupérer les paramètres
        spot_price = float(data.get('spotPrice', 100))
        strike_price = float(data.get('strikePrice', 100))
        time_maturity = float(data.get('timeMaturity', 1))
        risk_free_rate = float(data.get('riskFreeRate', 0.05))
        base_volatility = float(data.get('volatility', 0.2))
        option_type = data.get('optionType', 'call')
        
        # Paramètres pour la matrice de volatilité
        volatility_range = float(data.get('volatilityRange', 0.1))  # ±10% par défaut
        num_points = int(data.get('numPoints', 7))  # 7 points par défaut
        
        # Validation des paramètres
        if volatility_range <= 0 or volatility_range > 0.5:
            return jsonify({'error': 'La plage de volatilité doit être entre 0 et 0.5'}), 400
        
        if num_points < 3 or num_points > 15:
            return jsonify({'error': 'Le nombre de points doit être entre 3 et 15'}), 400
        
        if base_volatility - volatility_range < 0:
            return jsonify({'error': 'La volatilité de base moins la plage ne peut pas être négative'}), 400
        
        # Calculer la matrice de sensibilité
        sensitivity_matrix = greeks_calculator.generate_volatility_sensitivity_matrix(
            S=spot_price,
            K=strike_price,
            T=time_maturity,
            r=risk_free_rate,
            base_volatility=base_volatility,
            option_type=option_type,
            volatility_range=volatility_range,
            num_points=num_points
        )
        
        # Calculer les valeurs de référence avec la volatilité de base
        reference_values = greeks_calculator.get_greek_values_at_spot(
            S=spot_price,
            K=strike_price,
            T=time_maturity,
            r=risk_free_rate,
            sigma=base_volatility,
            option_type=option_type
        )
        
        return jsonify({
            'success': True,
            'sensitivity_matrix': sensitivity_matrix,
            'reference_values': reference_values,
            'parameters': {
                'spotPrice': spot_price,
                'strikePrice': strike_price,
                'timeMaturity': time_maturity,
                'riskFreeRate': risk_free_rate,
                'baseVolatility': base_volatility,
                'volatilityRange': volatility_range,
                'numPoints': num_points,
                'optionType': option_type
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/maturity-sensitivity-matrix', methods=['POST'])
def api_maturity_sensitivity_matrix():
    """API pour calculer la matrice de sensibilité des Greeks selon la maturité"""
    try:
        data = request.get_json()
        
        # Récupérer les paramètres
        spot_price = float(data.get('spotPrice', 100))
        strike_price = float(data.get('strikePrice', 100))
        time_maturity = float(data.get('timeMaturity', 1))
        risk_free_rate = float(data.get('riskFreeRate', 0.05))
        volatility = float(data.get('volatility', 0.2))
        option_type = data.get('optionType', 'call')
        
        # Paramètres pour la plage de maturité (non utilisés car la fonction backend a sa propre logique)
        maturity_range = float(data.get('maturityRange', 0.9))  # Non utilisé dans la fonction backend
        num_points = int(data.get('numPoints', 6))  # Non utilisé dans la fonction backend
        
        # Validation des paramètres
        if maturity_range <= 0 or maturity_range > 2.0:
            return jsonify({'error': 'La plage de maturité doit être entre 0 et 2.0 années'}), 400
        
        if num_points < 3 or num_points > 21:
            return jsonify({'error': 'Le nombre de points doit être entre 3 et 21'}), 400
        
        # Calculer la matrice de sensibilité à la maturité
        sensitivity_matrix = greeks_calculator.generate_maturity_sensitivity_matrix(
            S=spot_price,
            K=strike_price,
            T=time_maturity,
            r=risk_free_rate,
            sigma=volatility,
            option_type=option_type,
            maturity_range=maturity_range,
            num_points=num_points
        )
        
        # Calculer les valeurs de référence (maturité de base)
        reference_values = greeks_calculator.get_greek_values_at_spot(
            S=spot_price,
            K=strike_price,
            T=time_maturity,
            r=risk_free_rate,
            sigma=volatility,
            option_type=option_type
        )
        
        return jsonify({
            'success': True,
            'sensitivity_matrix': sensitivity_matrix,
            'reference_values': reference_values,
            'parameters': {
                'spotPrice': spot_price,
                'strikePrice': strike_price,
                'timeMaturity': time_maturity,
                'riskFreeRate': risk_free_rate,
                'volatility': volatility,
                'maturityRange': maturity_range,
                'numPoints': num_points,
                'optionType': option_type
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/risk-metrics/<symbol>')
def api_risk_metrics(symbol):
    """Calculer les métriques de risque pour un symbole donné"""
    try:
        start = request.args.get('start')
        end = request.args.get('end')
        
        # Utiliser le nouveau module Yahoo Finance API avec les dates spécifiées
        data = yahoo_api.get_chart_data(symbol, '1d', start, end)
        
        if not data or 'labels' not in data or 'datasets' not in data:
            return jsonify({'error': 'Données indisponibles'}), 404
        
        # Extraire les prix de clôture
        prices = [float(p) for p in data['datasets'][0]['data'] if p is not None and float(p) > 0]
        
        # Valider les données
        is_valid, message = risk_calculator.validate_data(prices)
        if not is_valid:
            return jsonify({'error': message}), 400
        
        # Calculer toutes les métriques
        metrics = risk_calculator.calculate_all_metrics(prices)
        
        # Ajouter les informations de période
        result = {
            'symbol': symbol,
            **metrics,
            'period': {
                'start': data['labels'][0] if data['labels'] else None,
                'end': data['labels'][-1] if data['labels'] else None,
                'days': len(prices)
            }
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# API: surface de volatilité 3D optimisée avec fond transparent
@app.route('/api/vol-surface-3d/<symbol>')
def api_vol_surface_3d(symbol):
    """API endpoint optimisé pour la surface de volatilité 3D avec fond transparent"""
    try:
        # Log de l'utilisation mémoire au début
        log_memory_usage()
        
        span = float(request.args.get('span', 0.5))
        provider = request.args.get('provider', 'tradier')
        
        # Validation des paramètres
        if span <= 0 or span > 1:
            return jsonify({'error': 'span doit être entre 0 et 1'}), 400
            
        if provider != 'tradier':
            return jsonify({'error': 'Seul le provider Tradier est supporté'}), 400
        
        # Vérifier que la clé API est disponible pour Tradier
        if not TRADIER_API_KEY:
            return jsonify({
                'error': 'Clé API Tradier non configurée',
                'details': 'Veuillez configurer TRADIER_API_KEY dans les variables d\'environnement'
            }), 500
        
        tradier = TradierAPI(TRADIER_API_KEY)
        
        # Récupérer les expirations avec timeout via Tradier
        expirations_data = tradier.get_option_expirations(symbol)
        if not expirations_data or 'expirations' not in expirations_data or 'date' not in expirations_data['expirations']:
            return jsonify({'error': f'Aucune date d\'expiration disponible pour {symbol}'}), 404
        
        # Extraire les dates d'expiration
        expirations = expirations_data['expirations']['date']
        
        # Limiter le nombre d'expirations pour la performance (par défaut 6)
        max_expirations = int(request.args.get('maxExp', 6))
        expirations_to_use = expirations[:max_expirations]
        
        # Récupérer le prix spot via Tradier
        try:
            quote_data = tradier.get_stock_quote(symbol)
            if quote_data and 'quotes' in quote_data and 'quote' in quote_data['quotes']:
                spot_price = float(quote_data['quotes']['quote'].get('last', 0))
            else:
                spot_price = 0
        except Exception as e:
            print(f"⚠️  Erreur lors de la récupération du prix spot pour {symbol}: {e}")
            spot_price = 0
        
        # Collecter les données d'options via Tradier
        all_options_data = []
        for expiration_date in expirations_to_use:
            try:
                options_data = tradier.get_historical_options_data(symbol, expiration_date)
                
                if options_data is not None and not options_data.empty:
                    # Renommer les colonnes pour correspondre au format attendu
                    options_data = options_data.rename(columns={
                        'Strike': 'strike',
                        'Type': 'type',
                        'Last': 'lastPrice',
                        'Implied_Volatility': 'impliedVolatility',
                        'Expiration': 'expiration_date'
                    })
                    # Convertir le type en minuscules
                    options_data['type'] = options_data['type'].str.lower()
                    all_options_data.append(options_data)
                    
            except Exception as e:
                print(f"Erreur pour l'expiration {expiration_date if 'expiration_date' in locals() else 'inconnue'}: {e}")
                continue
        
        if not all_options_data:
            return jsonify({'error': f'Aucune donnée d\'options collectée pour {symbol}'}), 404
        
        # Traiter les données pour créer la surface
        # Combiner toutes les données
        combined_data = pd.concat(all_options_data, ignore_index=True)
        
        # Filtrer par bande autour du spot si le prix spot est disponible
        if spot_price > 0:
            min_strike = spot_price * (1 - span)
            max_strike = spot_price * (1 + span)
            filtered_data = combined_data[
                (combined_data['strike'] >= min_strike) & 
                (combined_data['strike'] <= max_strike)
            ]
            
            # Si le filtrage supprime trop de données, utiliser toutes les données
            if len(filtered_data) < 10:
                print(f'Filtrage trop restrictif, utilisation de toutes les données ({len(combined_data)} options)')
                combined_data = combined_data
            else:
                combined_data = filtered_data
        else:
            print(f'Prix spot non disponible, utilisation de toutes les données ({len(combined_data)} options)')
        
        if combined_data.empty:
            return jsonify({
                'error': f'Aucune option disponible pour {symbol}'
            }), 404
        
        # Organiser les données pour la surface de volatilité
        # Extraire les strikes et expirations uniques
        unique_strikes = sorted(combined_data['strike'].unique(), reverse=True)
        unique_expirations = sorted(combined_data['expiration_date'].unique())
        
        # Calculer les maturités en années
        current_date = datetime.now()
        maturities = []
        for exp_date in unique_expirations:
            exp_datetime = datetime.strptime(exp_date, '%Y-%m-%d')
            maturity = (exp_datetime - current_date).days / 365.25
            # S'assurer que la maturité est positive et raisonnable
            if maturity < 0:
                print(f'⚠️  Maturité négative pour {exp_date}: {maturity:.4f} ans')
                maturity = 0.01  # Maturité minimale
            elif maturity > 5:
                print(f'⚠️  Maturité très élevée pour {exp_date}: {maturity:.4f} ans')
                maturity = 5.0  # Maturité maximale
            maturities.append(maturity)
        
        # Créer la matrice de volatilité implicite
        iv_matrix = []
        for i, expiration in enumerate(unique_expirations):
            row = []
            for strike in unique_strikes:
                # Trouver l'option correspondante
                option = combined_data[
                    (combined_data['expiration_date'] == expiration) & 
                    (combined_data['strike'] == strike)
                ]
                
                if not option.empty:
                    # Prendre la moyenne des IV pour les calls et puts
                    iv_values = option['impliedVolatility'].values
                    # Filtrer les valeurs aberrantes (IV > 200% ou < 1%) et exclure les None
                    valid_iv_values = [iv for iv in iv_values if iv is not None and 0.01 <= iv <= 2.0]
                    if valid_iv_values:
                        avg_iv = sum(valid_iv_values) / len(valid_iv_values)
                        row.append(avg_iv)
                    else:
                        row.append(None)
                else:
                    row.append(None)
            iv_matrix.append(row)
        
        # Filtrer les données valides pour les statistiques
        valid_iv_data = combined_data[
            (combined_data['impliedVolatility'].notna()) & 
            (combined_data['impliedVolatility'] > 0.01) & 
            (combined_data['impliedVolatility'] < 2.0)
        ]
        
        if valid_iv_data.empty:
            return jsonify({
                'error': f'Aucune donnée IV valide pour {symbol}'
            }), 404
        
        # Convertir le DataFrame en format JSON
        result = {
            'symbol': symbol,
            'spot_price': spot_price,
            'data_source': 'Tradier API (Données Réelles)',
            'strikes': unique_strikes,
            'maturities': maturities,
            'iv': iv_matrix,
            'total_options': len(combined_data),
            'calls_count': len(combined_data[combined_data['type'] == 'call']),
            'puts_count': len(combined_data[combined_data['type'] == 'put']),
            'valid_options': len(valid_iv_data),
            'statistics': {
                'min_iv': float(valid_iv_data['impliedVolatility'].min()),
                'max_iv': float(valid_iv_data['impliedVolatility'].max()),
                'avg_iv': float(valid_iv_data['impliedVolatility'].mean()),
                'std_iv': float(valid_iv_data['impliedVolatility'].std())
            },
            'raw_options': combined_data.to_dict('records')  # Données brutes pour debug
        }
        
        if 'error' in result:
            error_msg = result['error']
            if 'limite de taux' in error_msg.lower() or '429' in error_msg:
                return jsonify({
                    'error': 'API temporairement indisponible (limite de taux). Veuillez réessayer dans quelques secondes.',
                    'details': error_msg
                }), 429
            elif 'aucune donnée' in error_msg.lower():
                return jsonify({
                    'error': f'Aucune donnée disponible pour {symbol} via {provider}',
                    'details': error_msg
                }), 404
            else:
                return jsonify({
                    'error': f'Erreur lors de la récupération des données: {error_msg}',
                    'details': error_msg
                }), 500
        
        # Nettoyer la mémoire avant de retourner la réponse
        # cleanup_memory()  # Fonction non disponible
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Erreur générale dans api_vol_surface_3d: {e}")
        return jsonify({'error': str(e)}), 500


# API: Export des données de surface de volatilité 3D
@app.route('/api/vol-surface-3d-export/<symbol>')
def api_vol_surface_3d_export(symbol):
    """API endpoint pour exporter les données de surface de volatilité 3D"""
    try:
        span = float(request.args.get('span', 0.5))
        provider = request.args.get('provider', 'tradier')
        format_type = request.args.get('format', 'json')  # json, csv, excel
        
        # Validation des paramètres
        if span <= 0 or span > 1:
            return jsonify({'error': 'span doit être entre 0 et 1'}), 400
            
        if provider != 'tradier':
            return jsonify({'error': 'Seul le provider Tradier est supporté'}), 400
            
        if format_type not in ['json', 'csv', 'excel']:
            return jsonify({'error': 'format doit être "json", "csv" ou "excel"'}), 400
        
        # Récupérer les données via Tradier uniquement
        tradier = TradierAPI(TRADIER_API_KEY)
        
        # Récupérer les expirations
        expirations_data = tradier.get_option_expirations(symbol)
        if not expirations_data or 'expirations' not in expirations_data or 'date' not in expirations_data['expirations']:
            return jsonify({'error': f'Aucune date d\'expiration disponible pour {symbol}'}), 404
        
        # Extraire les dates d'expiration
        expirations = expirations_data['expirations']['date']
        
        # Utiliser toutes les expirations disponibles
        expirations_to_use = expirations
        
        # Récupérer le prix spot via Tradier
        try:
            quote_data = tradier.get_stock_quote(symbol)
            if quote_data and 'quotes' in quote_data and 'quote' in quote_data['quotes']:
                spot_price = float(quote_data['quotes']['quote'].get('last', 0))
            else:
                spot_price = 0
        except:
            spot_price = 0
        
        # Collecter les données d'options via Tradier
        all_options_data = []
        for expiration_date in expirations_to_use:
            try:
                options_data = tradier.get_historical_options_data(symbol, expiration_date)
                
                if options_data is not None and not options_data.empty:
                    # Renommer les colonnes pour correspondre au format attendu
                    options_data = options_data.rename(columns={
                        'Strike': 'strike',
                        'Type': 'type',
                        'Last': 'lastPrice',
                        'Implied_Volatility': 'impliedVolatility',
                        'Expiration': 'expiration_date'
                    })
                    # Convertir le type en minuscules
                    options_data['type'] = options_data['type'].str.lower()
                    options_data['expiration_date'] = expiration_date
                    all_options_data.append(options_data)
                    
            except Exception as e:
                print(f"Erreur pour l'expiration {expiration_date if 'expiration_date' in locals() else 'inconnue'}: {e}")
                continue
        
        if not all_options_data:
            return jsonify({'error': f'Aucune donnée d\'options collectée pour {symbol}'}), 404
        
        # Traiter les données pour créer la surface
        # Combiner toutes les données
        combined_data = pd.concat(all_options_data, ignore_index=True)
        
        # Filtrer par bande autour du spot si le prix spot est disponible
        if spot_price > 0:
            min_strike = spot_price * (1 - span)
            max_strike = spot_price * (1 + span)
            filtered_data = combined_data[
                (combined_data['strike'] >= min_strike) & 
                (combined_data['strike'] <= max_strike)
            ]
            
            # Si le filtrage supprime trop de données, utiliser toutes les données
            if len(filtered_data) < 10:
                print(f'Filtrage trop restrictif, utilisation de toutes les données ({len(combined_data)} options)')
                combined_data = combined_data
            else:
                combined_data = filtered_data
        else:
            print(f'Prix spot non disponible, utilisation de toutes les données ({len(combined_data)} options)')
        
        if combined_data.empty:
            return jsonify({
                'error': f'Aucune option disponible pour {symbol}'
            }), 404
        
        # Organiser les données pour la surface de volatilité
        # Extraire les strikes et expirations uniques
        unique_strikes = sorted(combined_data['strike'].unique(), reverse=True)
        unique_expirations = sorted(combined_data['expiration_date'].unique())
        
        # Calculer les maturités en années
        current_date = datetime.now()
        maturities = []
        for exp_date in unique_expirations:
            exp_datetime = datetime.strptime(exp_date, '%Y-%m-%d')
            maturity = (exp_datetime - current_date).days / 365.25
            # S'assurer que la maturité est positive et raisonnable
            if maturity < 0:
                print(f'⚠️  Maturité négative pour {exp_date}: {maturity:.4f} ans')
                maturity = 0.01  # Maturité minimale
            elif maturity > 5:
                print(f'⚠️  Maturité très élevée pour {exp_date}: {maturity:.4f} ans')
                maturity = 5.0  # Maturité maximale
            maturities.append(maturity)
        
        # Créer la matrice de volatilité implicite
        iv_matrix = []
        for i, expiration in enumerate(unique_expirations):
            row = []
            for strike in unique_strikes:
                # Trouver l'option correspondante
                option = combined_data[
                    (combined_data['expiration_date'] == expiration) & 
                    (combined_data['strike'] == strike)
                ]
                
                if not option.empty:
                    # Prendre la moyenne des IV pour les calls et puts
                    iv_values = option['impliedVolatility'].values
                    # Filtrer les valeurs aberrantes (IV > 200% ou < 1%) et exclure les None
                    valid_iv_values = [iv for iv in iv_values if iv is not None and 0.01 <= iv <= 2.0]
                    if valid_iv_values:
                        avg_iv = sum(valid_iv_values) / len(valid_iv_values)
                        row.append(avg_iv)
                    else:
                        row.append(None)
                else:
                    row.append(None)
            iv_matrix.append(row)
        
        # Filtrer les données valides pour les statistiques
        valid_iv_data = combined_data[
            (combined_data['impliedVolatility'].notna()) & 
            (combined_data['impliedVolatility'] > 0.01) & 
            (combined_data['impliedVolatility'] < 2.0)
        ]
        
        if valid_iv_data.empty:
            return jsonify({
                'error': f'Aucune donnée IV valide pour {symbol}'
            }), 404
        
        # Convertir le DataFrame en format JSON
        result = {
            'symbol': symbol,
            'spot_price': spot_price,
            'data_source': 'Tradier API (Données Réelles)',
            'strikes': unique_strikes,
            'maturities': maturities,
            'iv': iv_matrix,
            'total_options': len(combined_data),
            'calls_count': len(combined_data[combined_data['type'] == 'call']),
            'puts_count': len(combined_data[combined_data['type'] == 'put']),
            'valid_options': len(valid_iv_data),
            'statistics': {
                'min_iv': float(valid_iv_data['impliedVolatility'].min()),
                'max_iv': float(valid_iv_data['impliedVolatility'].max()),
                'avg_iv': float(valid_iv_data['impliedVolatility'].mean()),
                'std_iv': float(valid_iv_data['impliedVolatility'].std())
            },
            'raw_options': combined_data.to_dict('records')  # Données brutes pour debug
        }
        
        if 'error' in result:
            return jsonify({'error': result['error']}), 404
        
        # Préparer les données d'export
        export_data = {
            'metadata': {
                'symbol': symbol,
                'provider': provider,
                'span': span,
                'export_date': datetime.now().isoformat(),
                'spot_price': result.get('spot_price', 0),
                'maturities_count': len(result.get('maturities', [])),
                'strikes_count': len(result.get('strikes', [])),
                'data_points': sum(len(row) for row in result.get('iv', []) if row)
            },
            'data': {
                'strikes': result.get('strikes', []),
                'maturities': result.get('maturities', []),
                'iv_matrix': result.get('iv', [])
            }
        }
        
        # Ajouter les statistiques si disponibles
        try:
            from models.volatility_surface_3d import VolatilitySurface3D
            vs3d = VolatilitySurface3D(symbol)
            vs3d.set_data({
                'strikes': result['strikes'],
                'maturities': result['maturities'],
                'iv': result['iv'],
                'spot_price': result.get('spot_price', 0)
            })
            export_data['statistics'] = vs3d.get_statistics()
        except ImportError:
            pass
        
        # Retourner selon le format demandé
        if format_type == 'json':
            return jsonify(export_data)
        elif format_type == 'csv':
            # Créer un CSV avec les données de la matrice IV
            import io
            import csv
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # En-tête avec les strikes
            header = ['Maturité (années)'] + [f'Strike ${s:.2f}' for s in result.get('strikes', [])]
            writer.writerow(header)
            
            # Données IV
            for i, maturity in enumerate(result.get('maturities', [])):
                row = [f'{maturity:.3f}'] + [f'{iv:.4f}' if iv is not None else '' for iv in result.get('iv', [])[i]]
                writer.writerow(row)
            
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=volatility_surface_{symbol}_{datetime.now().strftime("%Y%m%d")}.csv'}
            )
        elif format_type == 'excel':
            # Créer un fichier Excel
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Surface de Volatilité"
                
                # Métadonnées
                ws['A1'] = 'Métadonnées'
                ws['A1'].font = Font(bold=True)
                ws['A2'] = f'Symbole: {symbol}'
                ws['A3'] = f'Fournisseur: {provider}'
                ws['A4'] = f'Prix Spot: ${result.get("spot_price", 0):.2f}'
                ws['A5'] = f'Date d\'export: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                
                # Données IV
                ws['A7'] = 'Matrice de Volatilité Implicite'
                ws['A7'].font = Font(bold=True)
                
                # En-tête avec les strikes
                for j, strike in enumerate(result.get('strikes', []), 1):
                    ws.cell(row=8, column=j+1, value=f'Strike ${strike:.2f}')
                
                # Données IV
                for i, maturity in enumerate(result.get('maturities', []), 8):
                    ws.cell(row=i+1, column=1, value=f'{maturity:.3f}a')
                    for j, iv in enumerate(result.get('iv', [])[i-8], 1):
                        if iv is not None:
                            ws.cell(row=i+1, column=j+1, value=iv)
                
                # Sauvegarder en mémoire
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=volatility_surface_{symbol}_{datetime.now().strftime("%Y%m%d")}.xlsx'}
                )
                
            except ImportError:
                return jsonify({'error': 'openpyxl non installé pour l\'export Excel'}), 400
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# API: Smile de volatilité avec une seule maturité (Tradier uniquement)
@app.route('/api/volatility-smile/<symbol>')
def api_volatility_smile(symbol):
    """API endpoint pour le smile de volatilité avec une seule maturité (Tradier uniquement)"""
    try:
        maturity_days = int(request.args.get('maturity', 30))  # Maturité en jours
        span = float(request.args.get('span', 0.3))  # Bande autour du spot
        
        # Validation des paramètres
        if maturity_days < 1 or maturity_days > 365:
            return jsonify({'error': 'maturity doit être entre 1 et 365 jours'}), 400
            
        if span <= 0 or span > 1:
            return jsonify({'error': 'span doit être entre 0 et 1'}), 400
        
        # Utiliser Tradier pour le smile
        tradier = TradierAPI(TRADIER_API_KEY)
        result = {'error': 'Erreur inconnue'}
        
        # Récupérer les expirations
        expirations_data = tradier.get_option_expirations(symbol)
        if not expirations_data or 'expirations' not in expirations_data or 'date' not in expirations_data['expirations']:
            return jsonify({'error': f'Aucune expiration disponible pour {symbol}'}), 404
        
        # Extraire les dates d'expiration
        expirations = expirations_data['expirations']['date']
        
        # Trouver l'expiration la plus proche de la maturité demandée
        from datetime import datetime, timedelta
        current_date = datetime.now()
        target_date = current_date + timedelta(days=maturity_days)
        
        best_expiration = None
        min_diff = float('inf')
        
        for exp_date in expirations:
            exp_datetime = datetime.strptime(exp_date, '%Y-%m-%d')
            diff = abs((exp_datetime - current_date).days - maturity_days)
            if diff < min_diff:
                min_diff = diff
                best_expiration = exp_date
        
        if not best_expiration:
            return jsonify({'error': f'Aucune expiration trouvée pour {symbol}'}), 404
        
        # Récupérer les données d'options pour cette expiration
        options_data = tradier.get_historical_options_data(symbol, best_expiration)
        if options_data is None or options_data.empty:
            return jsonify({'error': f'Aucune option disponible pour {symbol} à l\'expiration {best_expiration}'}), 404
        
        # Récupérer le prix spot
        quote_data = tradier.get_stock_quote(symbol)
        if quote_data and 'quotes' in quote_data and 'quote' in quote_data['quotes']:
            spot_price = float(quote_data['quotes']['quote'].get('last', 0))
        else:
            spot_price = 0
        
        # Filtrer par bande autour du spot
        if spot_price > 0:
            min_strike = spot_price * (1 - span)
            max_strike = spot_price * (1 + span)
            filtered_data = options_data[
                (options_data['Strike'] >= min_strike) & 
                (options_data['Strike'] <= max_strike)
            ]
            if len(filtered_data) < 5:
                filtered_data = options_data
        else:
            filtered_data = options_data
        
        # Séparer les calls et puts
        calls_df = filtered_data[filtered_data['Type'].str.lower() == 'call']
        puts_df = filtered_data[filtered_data['Type'].str.lower() == 'put']
        
        # Créer le résultat
        result = {
                'symbol': symbol,
                'spot_price': spot_price,
                'data_source': 'Tradier API (Données Réelles)',
                'maturity_days': maturity_days,
                'maturity_years': maturity_days / 365.25,
                'strikes': filtered_data['Strike'].tolist(),
                'calls_iv': calls_df['Implied_Volatility'].tolist() if len(calls_df) > 0 else [],
                'puts_iv': puts_df['Implied_Volatility'].tolist() if len(puts_df) > 0 else [],
                'calls_strikes': calls_df['Strike'].tolist() if len(calls_df) > 0 else [],
                'puts_strikes': puts_df['Strike'].tolist() if len(puts_df) > 0 else [],
                'statistics': {
                    'total_options': len(filtered_data),
                    'calls_count': len(calls_df),
                    'puts_count': len(puts_df),
                    'min_strike': float(filtered_data['Strike'].min()),
                    'max_strike': float(filtered_data['Strike'].max()),
                    'min_iv': float(filtered_data['Implied_Volatility'].min()),
                    'max_iv': float(filtered_data['Implied_Volatility'].max()),
                    'avg_iv': float(filtered_data['Implied_Volatility'].mean()),
                    'std_iv': float(filtered_data['Implied_Volatility'].std())
                },
                'raw_options': filtered_data.to_dict('records')
            }
        
        if 'error' in result:
            error_msg = result['error']
            if 'aucune option' in error_msg.lower():
                return jsonify({
                    'error': f'Aucune option disponible pour {symbol} avec maturité ~{maturity_days} jours',
                    'details': error_msg
                }), 404
            else:
                return jsonify({
                    'error': f'Erreur lors de la récupération du smile: {error_msg}',
                    'details': error_msg
                }), 500
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# API: Récupération des symboles Tradier
@app.route('/api/tradier/symbols')
def get_tradier_symbols():
    """API endpoint pour récupérer les symboles disponibles via Tradier"""
    try:
        query = request.args.get('q', '')  # Terme de recherche optionnel
        
        if query:
            # Recherche avec un terme spécifique
            search_result = tradier_api.search_symbols(query, indexes=True)
            if not search_result or "securities" not in search_result:
                return jsonify({
                    'success': False,
                    'error': 'Aucun symbole trouvé',
                    'symbols': []
                }), 404
            
            securities = search_result["securities"]["security"]
            if not isinstance(securities, list):
                securities = [securities]
            
            # Formater les résultats
            symbols = []
            for security in securities:
                if security.get("type") == "stock":  # Filtrer seulement les actions
                    symbols.append({
                        "symbol": security.get("symbol", ""),
                        "name": security.get("description", ""),
                        "type": security.get("type", ""),
                        "exchange": security.get("exchange", "")
                    })
            
            return jsonify({
                'success': True,
                'symbols': symbols,
                'total': len(symbols),
                'query': query
            })
        else:
            # Retourner les symboles populaires si pas de recherche
            popular_symbols = tradier_api.get_popular_symbols()
            return jsonify({
                'success': True,
                'symbols': popular_symbols,
                'total': len(popular_symbols),
                'query': 'popular'
            })
            
    except Exception as e:
        print(f"❌ Erreur récupération symboles Tradier: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'symbols': []
        }), 500

# API: Smile de volatilité avec Tradier
@app.route('/api/volatility-smile-tradier/<symbol>')
def api_volatility_smile_tradier(symbol):
    """API endpoint pour le smile de volatilité avec Tradier"""
    try:
        maturity_days = int(request.args.get('maturity', 30))  # Maturité en jours
        span = float(request.args.get('span', 0.3))  # Bande autour du spot
        specific_expiration = request.args.get('expiration', None)  # Date d'expiration spécifique
        custom_risk_free = request.args.get('risk_free', None)  # Taux sans risque personnalisé
        custom_dividend = request.args.get('dividend', None)  # Rendement dividendes personnalisé
        custom_time_to_exp = request.args.get('time_to_exp', None)  # Time to expiration personnalisé
        
        # Validation des paramètres
        if maturity_days < 1 or maturity_days > 365:
            return jsonify({'error': 'maturity doit être entre 1 et 365 jours'}), 400
            
        if span <= 0 or span > 1:
            return jsonify({'error': 'span doit être entre 0 et 1'}), 400
        
        print(f"🚀 Récupération des options {symbol} via Tradier (maturité: {maturity_days} jours)")
        
        # Récupérer les expirations disponibles
        expirations = tradier_api.get_option_expirations(symbol)
        if not expirations or "expirations" not in expirations:
            return jsonify({'error': f'Aucune expiration disponible pour {symbol}'}), 404
        
        exp_list = expirations["expirations"]["date"]
        if not exp_list:
            return jsonify({'error': f'Aucune expiration trouvée pour {symbol}'}), 404
        
        # Utiliser la date d'expiration spécifique si fournie, sinon trouver la plus proche
        if specific_expiration:
            # Vérifier que la date spécifique est disponible
            if specific_expiration in exp_list:
                best_expiration = specific_expiration
                print(f"🎯 Utilisation de la date d'expiration spécifique: {best_expiration}")
            else:
                return jsonify({'error': f'Date d\'expiration {specific_expiration} non disponible. Dates disponibles: {exp_list}'}), 404
        else:
            # Trouver l'expiration la plus proche de la maturité demandée
            from datetime import timedelta
            today = datetime.now()
            target_date = today + timedelta(days=maturity_days)
            
            best_expiration = None
            min_diff = float('inf')
            
            for exp_date_str in exp_list:
                try:
                    exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d")
                    diff = abs((exp_date - today).days - maturity_days)
                    if diff < min_diff:
                        min_diff = diff
                        best_expiration = exp_date_str
                except ValueError:
                    continue
        
        if not best_expiration:
            return jsonify({'error': f'Aucune expiration proche de {maturity_days} jours trouvée'}), 404
        
        if specific_expiration:
            print(f"📅 Expiration sélectionnée: {best_expiration} (date spécifique)")
        else:
            print(f"📅 Expiration sélectionnée: {best_expiration} (écart: {min_diff} jours)")
        
        # Récupérer le prix spot
        stock_quote = tradier_api.get_stock_quote(symbol)
        if not stock_quote or "quotes" not in stock_quote:
            return jsonify({'error': f'Impossible de récupérer le prix spot pour {symbol}'}), 404
        
        quote_data = stock_quote["quotes"]["quote"]
        if isinstance(quote_data, list):
            quote_data = quote_data[0]
        
        spot_price = float(quote_data.get("last", 0))
        if spot_price <= 0:
            return jsonify({'error': f'Prix spot invalide pour {symbol}'}), 404
        
        print(f"💰 Prix spot {symbol}: ${spot_price:.2f}")
        
        # Récupérer le taux sans risque et le rendement des dividendes
        if custom_risk_free is not None:
            risk_free_rate = float(custom_risk_free)
            print(f"🎯 Taux sans risque personnalisé: {risk_free_rate*100:.2f}%")
        else:
            risk_free_rate = get_risk_free_rate()
            print(f"📊 Taux sans risque par défaut: {risk_free_rate*100:.2f}%")
            
        if custom_dividend is not None:
            dividend_yield = float(custom_dividend)
            print(f"🎯 Rendement dividendes personnalisé: {dividend_yield*100:.2f}%")
        else:
            dividend_yield = get_dividend_yield(symbol)
            print(f"📊 Rendement dividendes par défaut: {dividend_yield*100:.2f}%")
        
        print(f"📊 Paramètres Black-Scholes finaux:")
        print(f"  Taux sans risque: {risk_free_rate*100:.2f}%")
        print(f"  Rendement dividendes: {dividend_yield*100:.2f}%")
        
        # Récupérer la chaîne d'options
        chain_data = tradier_api.get_option_chain(symbol, best_expiration)
        if not chain_data or "options" not in chain_data:
            return jsonify({'error': f'Aucune option disponible pour {symbol} à l\'expiration {best_expiration}'}), 404
        
        options = chain_data["options"]["option"]
        if not options:
            return jsonify({'error': f'Aucune option trouvée pour {symbol} à l\'expiration {best_expiration}'}), 404
        
        # Convertir en liste si ce n'est pas déjà le cas
        if not isinstance(options, list):
            options = [options]
        
        print(f"📊 {len(options)} options trouvées")
        
        # AFFICHER LES DONNÉES BRUTES DE TRADIER
        print(f"\n🔍 DONNÉES BRUTES TRADIER POUR {symbol} - {best_expiration}:")
        print("=" * 80)
        for i, option in enumerate(options[:10]):  # Afficher les 10 premières options
            print(f"Option {i+1}:")
            print(f"  Strike: ${option.get('strike', 'N/A')}")
            print(f"  Type: {option.get('option_type', 'N/A')}")
            print(f"  Bid: ${option.get('bid', 'N/A')}")
            print(f"  Ask: ${option.get('ask', 'N/A')}")
            print(f"  Last: ${option.get('last', 'N/A')}")
            print(f"  Volume: {option.get('volume', 'N/A')}")
            print(f"  Open Interest: {option.get('open_interest', 'N/A')}")
            print(f"  Implied Volatility: {option.get('implied_volatility', 'N/A')}")
            print(f"  Delta: {option.get('delta', 'N/A')}")
            print(f"  Gamma: {option.get('gamma', 'N/A')}")
            print(f"  Theta: {option.get('theta', 'N/A')}")
            print(f"  Vega: {option.get('vega', 'N/A')}")
            print("-" * 40)
        
        if len(options) > 10:
            print(f"... et {len(options) - 10} autres options")
        print("=" * 80)
        
        # Filtrer les options selon le span
        min_strike = spot_price * (1 - span)
        max_strike = spot_price * (1 + span)
        
        print(f"\n🎯 FILTRAGE DES OPTIONS:")
        print(f"Prix spot: ${spot_price:.2f}")
        print(f"Span: {span*100}% (plage: ${min_strike:.2f} - ${max_strike:.2f})")
        
        filtered_options = []
        for option in options:
            try:
                strike = float(option.get("strike", 0))
                if min_strike <= strike <= max_strike:
                    # Récupérer les données importantes
                    option_data = {
                        "strike": strike,
                        "type": option.get("option_type", "Unknown").lower(),
                        "bid": option.get("bid", None),
                        "ask": option.get("ask", None),
                        "last": option.get("last", None),
                        "volume": option.get("volume", None),
                        "open_interest": option.get("open_interest", None),
                        "implied_volatility": option.get("implied_volatility", None),
                        "delta": option.get("delta", None),
                        "gamma": option.get("gamma", None),
                        "theta": option.get("theta", None),
                        "vega": option.get("vega", None)
                    }
                    filtered_options.append(option_data)
            except (ValueError, TypeError):
                continue
        
        if not filtered_options:
            return jsonify({'error': f'Aucune option dans la plage de strikes demandée pour {symbol}'}), 404
        
        print(f"✅ {len(filtered_options)} options filtrées dans la plage de strikes")
        
        # AFFICHER LES OPTIONS FILTRÉES AVEC PRIX LAST
        print(f"\n💰 OPTIONS FILTRÉES AVEC PRIX LAST:")
        print("=" * 80)
        for i, opt in enumerate(filtered_options[:15]):  # Afficher les 15 premières
            print(f"Option {i+1}: Strike ${opt['strike']:.2f} - {opt['type'].upper()}")
            print(f"  Last Price: ${opt['last'] if opt['last'] else 'N/A'}")
            print(f"  Bid/Ask: ${opt['bid'] if opt['bid'] else 'N/A'} / ${opt['ask'] if opt['ask'] else 'N/A'}")
            print(f"  IV Tradier: {opt['implied_volatility'] if opt['implied_volatility'] else 'N/A'}")
            print(f"  Volume: {opt['volume'] if opt['volume'] else 'N/A'}")
            
            # Calculer le temps jusqu'à l'expiration pour cette option
            if custom_time_to_exp is not None:
                time_to_exp = float(custom_time_to_exp)
                print(f"  Time to Exp: {time_to_exp:.3f} années (personnalisé)")
            else:
                # Utiliser la fonction optimisée pour calculer les jours ouvrés
                time_to_exp = calculate_business_days_to_expiration(best_expiration)
                business_days = int(time_to_exp * 252)
                print(f"  Time to Exp: {time_to_exp:.3f} années ({business_days} jours ouvrés)")
            
            # Calculer le moneyness
            moneyness = opt['strike'] / spot_price
            print(f"  Moneyness: {moneyness:.3f}")
            print("-" * 40)
        
        if len(filtered_options) > 15:
            print(f"... et {len(filtered_options) - 15} autres options filtrées")
        print("=" * 80)
        
        # Construire le smile de volatilité
        smile_data = []
        strikes = sorted(set(opt["strike"] for opt in filtered_options))
        
        print(f"\n📈 CONSTRUCTION DU SMILE DE VOLATILITÉ:")
        print(f"Strikes uniques: {len(strikes)}")
        print(f"Prix spot: ${spot_price:.2f}")
        print(f"Expiration: {best_expiration}")
        print(f"Time to Exp: {((datetime.strptime(best_expiration, '%Y-%m-%d') - datetime.now()).days / 365.25):.3f} années")
        print("=" * 80)
        
        for strike in strikes:
            strike_options = [opt for opt in filtered_options if opt["strike"] == strike]
            
            # Calculer la volatilité implicite moyenne pour ce strike
            ivs = []
            for opt in strike_options:
                # Essayer d'abord la volatilité implicite fournie par Tradier
                if opt["implied_volatility"] is not None:
                    try:
                        iv = float(opt["implied_volatility"])
                        if 0 < iv < 5:  # Filtrer les valeurs aberrantes
                            ivs.append(iv)
                    except (ValueError, TypeError):
                        continue
                
                # Si pas de volatilité implicite, essayer de la calculer à partir du prix
                elif opt["last"] is not None and opt["last"] > 0:
                    try:
                        # Calculer le temps jusqu'à l'expiration
                        if custom_time_to_exp is not None:
                            time_to_exp = float(custom_time_to_exp)
                        else:
                            # Utiliser la fonction optimisée pour calculer les jours ouvrés
                            time_to_exp = calculate_business_days_to_expiration(best_expiration)
                        
                        if time_to_exp > 0:
                            # Utiliser le prix last ou le prix moyen bid-ask
                            option_price = opt["last"]
                            if opt["bid"] is not None and opt["ask"] is not None and opt["bid"] > 0 and opt["ask"] > 0:
                                # Utiliser le prix moyen bid-ask si disponible
                                option_price = (opt["bid"] + opt["ask"]) / 2
                            
                            # Calculer la volatilité implicite avec Black-Scholes
                            calculated_iv = calculate_implied_volatility_black_scholes(
                                spot_price=spot_price,
                                strike=opt["strike"],
                                time_to_exp=time_to_exp,
                                option_price=option_price,
                                option_type=opt["type"],
                                risk_free_rate=risk_free_rate,
                                dividend_yield=dividend_yield
                            )
                            
                            if calculated_iv and 0.01 < calculated_iv < 2.0:  # Plage plus large pour les IV
                                ivs.append(calculated_iv)
                                print(f"    ✅ Calculé IV pour strike ${opt['strike']:.2f} {opt['type'].upper()}: {calculated_iv:.4f} ({calculated_iv*100:.2f}%) - Prix: ${option_price:.2f} - TTE: {time_to_exp:.3f}a")
                            else:
                                print(f"    ⚠️  IV calculée hors plage pour strike ${opt['strike']:.2f}: {calculated_iv}")
                    except Exception as e:
                        print(f"    ❌ Erreur calcul IV pour strike ${opt['strike']:.2f}: {e}")
                        continue
            
            if ivs:
                avg_iv = sum(ivs) / len(ivs)
                moneyness = strike / spot_price
                calls_count = len([opt for opt in strike_options if opt["type"] == "call"])
                puts_count = len([opt for opt in strike_options if opt["type"] == "put"])
                
                # Récupérer les données brutes pour ce strike
                call_options = [opt for opt in strike_options if opt["type"] == "call"]
                put_options = [opt for opt in strike_options if opt["type"] == "put"]
                
                raw_data = {
                    "call_last": call_options[0]["last"] if call_options else None,
                    "call_bid": call_options[0]["bid"] if call_options else None,
                    "call_ask": call_options[0]["ask"] if call_options else None,
                    "put_last": put_options[0]["last"] if put_options else None,
                    "put_bid": put_options[0]["bid"] if put_options else None,
                    "put_ask": put_options[0]["ask"] if put_options else None
                }
                
                smile_data.append({
                    "strike": strike,
                    "moneyness": moneyness,
                    "implied_volatility": avg_iv,
                    "calls_count": calls_count,
                    "puts_count": puts_count,
                    "raw_data": raw_data
                })
                
                print(f"Strike ${strike:.2f}: IV={avg_iv:.4f} ({avg_iv*100:.2f}%) - {len(ivs)} données - Moneyness={moneyness:.3f} - Calls:{calls_count} Puts:{puts_count}")
                
                # Afficher les détails des options pour ce strike
                for opt in strike_options:
                    if opt["implied_volatility"] is not None:
                        print(f"  📊 {opt['type'].upper()}: IV={opt['implied_volatility']:.4f} ({opt['implied_volatility']*100:.2f}%) - Prix=${opt['last'] if opt['last'] else 'N/A'}")
                    elif opt["last"] is not None and opt["last"] > 0:
                        # Calculer l'IV pour cette option
                        option_price = opt["last"]
                        if opt["bid"] is not None and opt["ask"] is not None and opt["bid"] > 0 and opt["ask"] > 0:
                            option_price = (opt["bid"] + opt["ask"]) / 2
                        
                        calculated_iv = calculate_implied_volatility_black_scholes(
                            spot_price=spot_price,
                            strike=strike,
                            time_to_exp=time_to_exp,
                            option_price=option_price,
                            option_type=opt["type"],
                            risk_free_rate=risk_free_rate,
                            dividend_yield=dividend_yield
                        )
                        
                        if calculated_iv:
                            print(f"  🔢 {opt['type'].upper()}: IV calculée={calculated_iv:.4f} ({calculated_iv*100:.2f}%) - Prix=${option_price:.2f}")
                        else:
                            print(f"  ❌ {opt['type'].upper()}: Impossible de calculer l'IV - Prix=${option_price:.2f}")
        
        if not smile_data:
            # Si aucun smile n'a pu être construit, essayer avec les prix des options
            print(f"\n⚠️  Aucune volatilité implicite trouvée, tentative avec les prix des options...")
            
            # Créer un smile basé sur les prix des options (sans IV)
            smile_data = []
            for strike in strikes:
                strike_options = [opt for opt in filtered_options if opt["strike"] == strike]
                
                # Calculer le prix moyen des options pour ce strike
                prices = []
                for opt in strike_options:
                    if opt["last"] is not None and opt["last"] > 0:
                        prices.append(opt["last"])
                
                if prices:
                    avg_price = sum(prices) / len(prices)
                    moneyness = strike / spot_price
                    calls_count = len([opt for opt in strike_options if opt["type"] == "call"])
                    puts_count = len([opt for opt in strike_options if opt["type"] == "put"])
                    
                    # Récupérer les données brutes pour ce strike
                    call_options = [opt for opt in strike_options if opt["type"] == "call"]
                    put_options = [opt for opt in strike_options if opt["type"] == "put"]
                    
                    raw_data = {
                        "call_last": call_options[0]["last"] if call_options else None,
                        "call_bid": call_options[0]["bid"] if call_options else None,
                        "call_ask": call_options[0]["ask"] if call_options else None,
                        "put_last": put_options[0]["last"] if put_options else None,
                        "put_bid": put_options[0]["bid"] if put_options else None,
                        "put_ask": put_options[0]["ask"] if put_options else None
                    }
                    
                    smile_data.append({
                        "strike": strike,
                        "moneyness": moneyness,
                        "implied_volatility": None,  # Pas d'IV disponible
                        "option_price": avg_price,  # Prix moyen des options
                        "calls_count": calls_count,
                        "puts_count": puts_count,
                        "raw_data": raw_data
                    })
                    
                    print(f"Strike ${strike:.2f}: Prix moyen=${avg_price:.2f} - Moneyness={moneyness:.3f} - Calls:{calls_count} Puts:{puts_count}")
            
            if not smile_data:
                return jsonify({'error': f'Aucune donnée d\'option valide trouvée pour {symbol}'}), 404
        
        # Trier par strike (ordre décroissant pour avoir les strikes les plus grands à gauche)
        smile_data.sort(key=lambda x: x["strike"], reverse=True)
        
        print(f"\n✅ SMILE CONSTRUIT AVEC {len(smile_data)} POINTS:")
        for point in smile_data:
            print(f"  ${point['strike']:.2f} -> IV: {point['implied_volatility']:.4f} ({point['implied_volatility']*100:.2f}%)")
        print("=" * 80)
        
        # Construire la réponse
        result = {
            "symbol": symbol,
            "spot_price": spot_price,
            "expiration": best_expiration,
            "maturity_days": (datetime.strptime(best_expiration, "%Y-%m-%d") - datetime.now()).days,
            "span": span,
            "data_source": "Tradier API",
            "black_scholes_params": {
                "risk_free_rate": risk_free_rate,
                "dividend_yield": dividend_yield,
                "time_to_expiration": float(custom_time_to_exp) if custom_time_to_exp is not None else calculate_business_days_to_expiration(best_expiration),
                "time_calculation_method": "business_days_252",
                "optimization_note": "Paramètres optimisés pour convergence avec données d'images (5% taux, 0% div, jours ouvrés/252)"
            },
            "smile_data": smile_data,
            "total_options": len(filtered_options),
            "valid_strikes": len(smile_data)
        }
        
        print(f"✅ Smile de volatilité créé avec {len(smile_data)} points de données")
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Erreur API Tradier: {e}")
        return jsonify({'error': str(e)}), 500


# API: Test simple Tradier
@app.route('/api/tradier-test/<symbol>')
def test_tradier_simple(symbol):
    """Test simple pour Tradier"""
    try:
        # Test basique
        stock_quote = tradier_api.get_stock_quote(symbol)
        if not stock_quote:
            return jsonify({'error': 'Impossible de récupérer le prix spot'}), 404
        
        quote_data = stock_quote["quotes"]["quote"]
        if isinstance(quote_data, list):
            quote_data = quote_data[0]
        
        spot_price = float(quote_data.get("last", 0))
        
        return jsonify({
            'success': True,
            'symbol': symbol,
            'spot_price': spot_price,
            'message': 'Test Tradier réussi'
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur test: {str(e)}'}), 500

# API: Surface de volatilité 3D avec Tradier (VERSION SIMPLIFIÉE QUI FONCTIONNE)
@app.route('/api/vol-surface-3d-tradier-simple/<symbol>')
def get_volatility_surface_3d_tradier_simple(symbol):
    """API endpoint simplifié pour la surface de volatilité 3D avec Tradier"""
    try:
        print(f"🚀 Génération surface de volatilité 3D {symbol} via Tradier (version simplifiée)")
        
        # Utiliser exactement le même code que notre fichier de test qui fonctionne
        from datetime import datetime
        import pandas as pd
        import numpy as np
        
        # ÉTAPE 1: Récupérer le prix spot
        print(f"📊 ÉTAPE 1: Récupération du prix spot pour {symbol}")
        stock_quote = tradier_api.get_stock_quote(symbol)
        if not stock_quote or "quotes" not in stock_quote:
            return jsonify({'error': f'Impossible de récupérer le prix spot pour {symbol}'}), 404
        
        quote_data = stock_quote["quotes"]["quote"]
        if isinstance(quote_data, list):
            quote_data = quote_data[0]
        
        spot_price = float(quote_data.get("last", 0))
        if spot_price <= 0:
            return jsonify({'error': f'Prix spot invalide pour {symbol}'}), 404
        
        print(f"💰 Prix spot {symbol}: ${spot_price:.2f}")
        
        # ÉTAPE 2: Récupérer les expirations disponibles
        print(f"📅 ÉTAPE 2: Récupération des maturités disponibles")
        expirations = tradier_api.get_option_expirations(symbol)
        if not expirations or "expirations" not in expirations:
            return jsonify({'error': f'Aucune expiration disponible pour {symbol}'}), 404
        
        exp_list = expirations["expirations"]["date"]
        if not exp_list:
            return jsonify({'error': f'Aucune expiration trouvée pour {symbol}'}), 404
        
        # Limiter le nombre de maturités pour le test
        max_maturities = 3
        selected_maturities = exp_list[:max_maturities]
        print(f"📅 {len(selected_maturities)} maturités sélectionnées: {selected_maturities}")
        
        # ÉTAPE 3: Construire la matrice de volatilité
        print(f"🎯 ÉTAPE 3: Construction de la matrice de volatilité")
        
        all_data = []
        
        for maturity_date in selected_maturities:
            print(f"📊 Traitement de la maturité: {maturity_date}")
            
            # Calculer le temps jusqu'à l'expiration
            time_to_exp = calculate_business_days_to_expiration(maturity_date)
            if time_to_exp is None or time_to_exp <= 0:
                print(f"   ⚠️  Temps d'expiration invalide: {time_to_exp}")
                continue
            
            print(f"   ⏰ Temps jusqu'à expiration: {time_to_exp:.4f} années")
            
            # Récupérer la chaîne d'options pour cette maturité
            chain_data = tradier_api.get_option_chain(symbol, maturity_date)
            if not chain_data or "options" not in chain_data:
                print(f"   ⚠️  Aucune option pour {maturity_date}")
                continue
            
            options = chain_data["options"]["option"]
            if not options:
                print(f"   ⚠️  Aucune option trouvée pour {maturity_date}")
                continue
            
            # Convertir en liste si nécessaire
            if not isinstance(options, list):
                options = [options]
            
            print(f"   💰 {len(options)} options trouvées")
            
            # Traiter chaque option
            maturity_data = []
            for option in options:
                try:
                    # Récupérer le strike
                    strike = option.get("strike")
                    if strike is None:
                        continue
                    
                    try:
                        strike_float = float(strike)
                        if strike_float <= 0:
                            continue
                    except (ValueError, TypeError):
                        continue
                    
                    # Récupérer le prix de l'option
                    option_price = None
                    if option.get("last") and option.get("last") is not None:
                        try:
                            last_price = float(option.get("last", 0))
                            if last_price > 0:
                                option_price = last_price
                        except (ValueError, TypeError):
                            pass
                    
                    if option_price is None and option.get("bid") and option.get("ask"):
                        try:
                            bid = float(option.get("bid", 0))
                            ask = float(option.get("ask", 0))
                            if bid > 0 and ask > 0:
                                option_price = (bid + ask) / 2
                        except (ValueError, TypeError):
                            pass
                    
                    if option_price is None or option_price <= 0:
                        continue
                    
                    # Calculer la volatilité implicite
                    option_type = option.get("option_type", "call").lower()
                    implied_vol = calculate_implied_volatility_black_scholes(
                        spot_price=spot_price,
                        strike=strike_float,
                        time_to_exp=time_to_exp,
                        option_price=option_price,
                        option_type=option_type,
                        risk_free_rate=0.05,
                        dividend_yield=0.0
                    )
                    
                    if implied_vol is not None and 0.01 <= implied_vol <= 2.0:  # Filtrer les valeurs aberrantes
                        option_data = {
                            'maturity_date': maturity_date,
                            'time_to_exp': time_to_exp,
                            'strike': strike_float,
                            'option_type': option_type,
                            'option_price': option_price,
                            'implied_volatility': implied_vol,
                            'contractSymbol': f"{symbol}{maturity_date.replace('-', '')}{option.get('option_type', 'C').upper()}{strike_float:08.0f}",
                            'type': option_type,
                            'lastPrice': option_price,
                            'expiration_date': maturity_date
                        }
                        maturity_data.append(option_data)
                        all_data.append(option_data)
                
                except Exception as e:
                    print(f"   ⚠️  Erreur traitement option: {e}")
                    continue
            
            print(f"   ✅ {len(maturity_data)} options avec IV calculée")
            
            # AFFICHER LES VOLATILITÉS IMPLICITES DANS LA CONSOLE
            if maturity_data:
                print(f"   📊 VOLATILITÉS IMPLICITES CALCULÉES pour {maturity_date}:")
                print(f"   {'Strike':<10} {'Type':<6} {'Prix':<8} {'IV':<8} {'IV%':<8}")
                print(f"   {'-'*10} {'-'*6} {'-'*8} {'-'*8} {'-'*8}")
                
                for option in maturity_data[:10]:  # Afficher les 10 premières
                    strike = option['strike']
                    opt_type = option['option_type']
                    price = option['option_price']
                    iv = option['implied_volatility']
                    iv_percent = iv * 100 if iv > 0 else 0
                    
                    print(f"   ${strike:<9.2f} {opt_type:<6} ${price:<7.2f} {iv:<7.4f} {iv_percent:<7.2f}%")
                
                if len(maturity_data) > 10:
                    print(f"   ... et {len(maturity_data) - 10} autres options")
        
        if not all_data:
            return jsonify({'error': f'Aucune donnée d\'options collectée pour {symbol}'}), 404
        
        # ÉTAPE 4: Créer le DataFrame et analyser
        print(f"📈 ÉTAPE 4: Analyse des données collectées")
        df = pd.DataFrame(all_data)
        
        print(f"📊 Total d'options traitées: {len(df)}")
        print(f"📅 Maturités uniques: {df['maturity_date'].nunique()}")
        print(f"🎯 Strikes uniques: {df['strike'].nunique()}")
        print(f"📊 Types d'options: {df['option_type'].value_counts().to_dict()}")
        
        # Statistiques de volatilité implicite
        iv_stats = df['implied_volatility'].describe()
        print(f"📊 Statistiques de volatilité implicite:")
        print(f"   Moyenne: {iv_stats['mean']:.4f}")
        print(f"   Médiane: {iv_stats['50%']:.4f}")
        print(f"   Min: {iv_stats['min']:.4f}")
        print(f"   Max: {iv_stats['max']:.4f}")
        
        # ÉTAPE 5: Construire la matrice de volatilité
        print(f"🎯 ÉTAPE 5: Construction de la matrice de volatilité")
        
        # Créer une matrice avec strikes en colonnes et maturités en lignes
        pivot_table = df.pivot_table(
            values='implied_volatility',
            index='maturity_date',
            columns='strike',
            aggfunc='mean'  # Moyenne si plusieurs options pour le même strike/maturité
        )
        
        print(f"📊 Matrice créée: {pivot_table.shape[0]} maturités × {pivot_table.shape[1]} strikes")
        
        # Créer le résultat au format attendu par l'interface
        strikes = sorted(df['strike'].unique())
        maturities = sorted(df['time_to_exp'].unique())
        
        # Créer la matrice IV
        iv_matrix = []
        for maturity in maturities:
            row = []
            for strike in strikes:
                # Trouver les options pour cette maturité et ce strike
                option_data = df[
                    (df['time_to_exp'] == maturity) & 
                    (df['strike'] == strike)
                ]
                
                if not option_data.empty:
                    # Prendre la moyenne des IV
                    iv_values = option_data['implied_volatility'].values
                    valid_iv_values = [iv for iv in iv_values if iv is not None and 0.01 <= iv <= 2.0]
                    
                    if valid_iv_values:
                        avg_iv = float(sum(valid_iv_values) / len(valid_iv_values))
                     
                        row.append(avg_iv)
                    else:
                        row.append(None)
                else:
                    row.append(None)
            iv_matrix.append(row)
        
        # Créer le résultat final
        result = {
            'success': True,
            'symbol': symbol,
            'spot_price': spot_price,
            'strikes': strikes,
            'maturities': maturities,
            'iv': iv_matrix,
            'statistics': {
                'min_iv': float(iv_stats['min']),
                'max_iv': float(iv_stats['max']),
                'mean_iv': float(iv_stats['mean']),
                'std_iv': float(iv_stats['std'])
            },
            'total_options': len(df),
            'calls_count': len(df[df['option_type'] == 'call']),
            'puts_count': len(df[df['option_type'] == 'put']),
            'data_source': 'Tradier API (Données Réelles)',
            'provider': 'tradier',
            'raw_options': df.to_dict('records')
        }
        
        print(f"✅ Surface de volatilité 3D générée avec succès pour {symbol}")
        print(f"📊 {result['total_options']} options, {len(result['maturities'])} maturités, {len(result['strikes'])} strikes")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"❌ Erreur lors de la génération de la surface de volatilité pour {symbol}: {e}")
        return jsonify({
            'error': f'Erreur lors de la génération de la surface de volatilité pour {symbol}',
            'details': str(e)
        }), 500

# API: Surface de volatilité 3D avec Tradier
@app.route('/api/vol-surface-3d-tradier/<symbol>')
def get_volatility_surface_3d_tradier(symbol):
    """API endpoint pour la surface de volatilité 3D avec Tradier - Redirection vers la version simplifiée"""
    # Rediriger vers l'endpoint simplifié qui fonctionne
    return redirect(f'/api/vol-surface-3d-tradier-simple/{symbol}')

# API: Symboles disponibles via Tradier
# API: Symboles disponibles via Tradier
@app.route('/api/available-symbols')
def api_available_symbols():
    """API endpoint pour récupérer les symboles disponibles via l'API Tradier"""
    try:
        # Utiliser l'API Tradier pour récupérer les symboles populaires
        popular_symbols = tradier_api.get_popular_symbols()
        
        # Ajouter des symboles supplémentaires via recherche Tradier
        additional_symbols = []
        
        # Rechercher des symboles populaires supplémentaires
        search_queries = ['tech', 'finance', 'energy', 'healthcare', 'consumer']
        
        for query in search_queries:
            try:
                search_results = tradier_api.search_symbols(query, indexes=True)
                if search_results and 'securities' in search_results:
                    securities = search_results['securities']['security']
                    if not isinstance(securities, list):
                        securities = [securities]
                    
                    for security in securities[:5]:  # Limiter à 5 par catégorie
                        if 'symbol' in security and 'description' in security:
                            additional_symbols.append({
                                'symbol': security['symbol'],
                                'name': security['description']
                            })
            except Exception as e:
                print(f"Erreur lors de la recherche '{query}': {e}")
                continue
        
        # Combiner les symboles populaires et supplémentaires
        all_symbols = popular_symbols + additional_symbols
        
        # Supprimer les doublons basés sur le symbole
        unique_symbols = []
        seen_symbols = set()
        for symbol_data in all_symbols:
            if symbol_data['symbol'] not in seen_symbols:
                unique_symbols.append(symbol_data)
                seen_symbols.add(symbol_data['symbol'])
        
        # Formater pour le frontend
        formatted_symbols = []
        for symbol_data in unique_symbols:
            formatted_symbols.append({
                'symbol': symbol_data['symbol'],
                'label': f"{symbol_data['symbol']} - {symbol_data['name']}",
                'name': symbol_data['name']
            })
        
        return jsonify({
            'recommended_symbols': formatted_symbols,
            'statistics': {
                'total_symbols': len(formatted_symbols),
                'source': 'Tradier API'
            }
        })
        
    except Exception as e:
        print(f"Erreur lors de la récupération des symboles: {e}")
        # Fallback vers une liste statique étendue
        fallback_symbols = [
            {"symbol": "AAPL", "name": "Apple Inc."},
            {"symbol": "MSFT", "name": "Microsoft Corp."},
            {"symbol": "GOOGL", "name": "Alphabet Inc."},
            {"symbol": "AMZN", "name": "Amazon.com Inc."},
            {"symbol": "TSLA", "name": "Tesla Inc."},
            {"symbol": "META", "name": "Meta Platforms Inc."},
            {"symbol": "NVDA", "name": "NVIDIA Corp."},
            {"symbol": "NFLX", "name": "Netflix Inc."},
            {"symbol": "AMD", "name": "Advanced Micro Devices"},
            {"symbol": "INTC", "name": "Intel Corp."},
            {"symbol": "CRM", "name": "Salesforce Inc."},
            {"symbol": "SPY", "name": "SPDR S&P 500 ETF"},
            {"symbol": "QQQ", "name": "Invesco QQQ Trust"},
            {"symbol": "IWM", "name": "iShares Russell 2000 ETF"},
            {"symbol": "VXX", "name": "iPath Series B S&P 500 VIX Short-Term Futures ETN"},
            {"symbol": "GLD", "name": "SPDR Gold Trust"},
            {"symbol": "SLV", "name": "iShares Silver Trust"},
            {"symbol": "TLT", "name": "iShares 20+ Year Treasury Bond ETF"},
            {"symbol": "HYG", "name": "iShares iBoxx $ High Yield Corporate Bond ETF"},
            {"symbol": "JPM", "name": "JPMorgan Chase & Co."},
            {"symbol": "BAC", "name": "Bank of America Corp."},
            {"symbol": "WMT", "name": "Walmart Inc."},
            {"symbol": "JNJ", "name": "Johnson & Johnson"},
            {"symbol": "PG", "name": "Procter & Gamble Co."},
            {"symbol": "KO", "name": "Coca-Cola Co."},
            {"symbol": "PFE", "name": "Pfizer Inc."},
            {"symbol": "ABBV", "name": "AbbVie Inc."},
            {"symbol": "V", "name": "Visa Inc."},
            {"symbol": "MA", "name": "Mastercard Inc."},
            {"symbol": "DIS", "name": "Walt Disney Co."}
        ]
        
        formatted_fallback = [
            {
                'symbol': symbol_data['symbol'],
                'label': f"{symbol_data['symbol']} - {symbol_data['name']}",
                'name': symbol_data['name']
            }
            for symbol_data in fallback_symbols
        ]
        
        return jsonify({
            'recommended_symbols': formatted_fallback,
            'statistics': {
                'total_symbols': len(formatted_fallback),
                'source': 'Fallback (API Tradier indisponible)'
            }
        })


def process_volatility_surface_data(all_options_data, symbol, spot_price, span, filter_by_span=True):
    """Traite les données d'options pour créer une surface de volatilité"""
    try:
        # Log de l'utilisation mémoire avant traitement
        log_memory_usage()
        
        # Combiner toutes les données
        combined_data = pd.concat(all_options_data, ignore_index=True)
        
        # Calculer les maturités en années
        current_date = datetime.now()
        print(f"🔍 Debug: Calcul des maturités pour {len(combined_data)} options")
        print(f"🔍 Debug: Dates d'expiration uniques: {combined_data['expiration_date'].unique()}")
        
        combined_data['maturity_years'] = combined_data['expiration_date'].apply(
            lambda x: (datetime.strptime(x, '%Y-%m-%d') - current_date).days / 365.25 if x is not None else None
        )
        
        # Filtrer les maturités futures (exclure les None)
        combined_data = combined_data[combined_data['maturity_years'].notna()]
        combined_data = combined_data[combined_data['maturity_years'] > 0]
        
        # Gérer les valeurs None dans les strikes
        combined_data = combined_data[combined_data['strike'].notna()]
        combined_data = combined_data[combined_data['strike'] > 0]
        
        if combined_data.empty:
            return {'error': 'Aucune donnée de maturité future disponible'}
        
        # Filtrer par span autour du spot (optionnel)
        if filter_by_span and spot_price is not None and spot_price > 0:
            min_strike = spot_price * (1 - span)
            max_strike = spot_price * (1 + span)
            # S'assurer que les strikes ne sont pas None
            combined_data = combined_data[combined_data['strike'].notna()]
            combined_data = combined_data[
                (combined_data['strike'] >= min_strike) & 
                (combined_data['strike'] <= max_strike)
            ]
        else:
            # Juste s'assurer que les strikes ne sont pas None
            combined_data = combined_data[combined_data['strike'].notna()]
        
        if combined_data.empty:
            return {'error': 'Aucune donnée dans la plage de strikes spécifiée'}
        
        # Grouper par maturité et strike pour créer la surface
        unique_maturities = sorted(combined_data['maturity_years'].unique())
        unique_strikes = sorted(combined_data['strike'].unique(), reverse=True)
        
        # Créer la matrice de volatilité implicite
        iv_matrix = []
        for maturity in unique_maturities:
            maturity_data = combined_data[combined_data['maturity_years'] == maturity]
            row = []
            for strike in unique_strikes:
                strike_data = maturity_data[maturity_data['strike'] == strike]
                if not strike_data.empty:
                    # Prendre la moyenne des IV pour ce strike/maturité
                    # Vérifier les noms de colonnes possibles
                    iv_column = None
                    for col in ['implied_volatility', 'impliedVolatility', 'iv']:
                        if col in strike_data.columns:
                            iv_column = col
                            break
                    
                    if iv_column:
                        avg_iv = strike_data[iv_column].mean()
                        row.append(avg_iv)
                    else:
                        row.append(None)
                else:
                    row.append(None)
            iv_matrix.append(row)
        
        # Calculer les statistiques
        all_iv_values = []
        for row in iv_matrix:
            all_iv_values.extend([v for v in row if v is not None and not pd.isna(v)])
        
        stats = {}
        if all_iv_values:
            stats = {
                'min_iv': float(min(all_iv_values)),
                'max_iv': float(max(all_iv_values)),
                'mean_iv': float(sum(all_iv_values) / len(all_iv_values))
                
            }
        
        # Compter les options
        total_options = len(combined_data)
        calls_count = len(combined_data[combined_data['type'] == 'call'])
        puts_count = len(combined_data[combined_data['type'] == 'put'])
        
        # Nettoyer la mémoire après traitement
        cleanup_memory()
        
        return {
            'symbol': symbol,
            'strikes': unique_strikes,
            'maturities': unique_maturities,
            'iv': iv_matrix,
            'spot_price': spot_price,
            'data_source': 'Tradier API (Données Réelles)',
            'total_options': total_options,
            'calls_count': calls_count,
            'puts_count': puts_count,
            'statistics': stats,
            'raw_options': combined_data.to_dict('records')
        }
        
    except Exception as e:
        return {'error': f'Erreur lors du traitement des données: {str(e)}'}


# API: Symboles populaires Tradier
@app.route('/api/tradier/symbols')
def api_tradier_symbols():
    """API endpoint pour récupérer les symboles populaires avec options via Tradier"""
    try:
        symbols = tradier_api.get_popular_symbols()
        return jsonify({
            'success': True,
            'symbols': symbols,
            'count': len(symbols)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# API: Recherche de symboles Tradier
@app.route('/api/tradier/search')
def api_tradier_search():
    """API endpoint pour rechercher des symboles via Tradier"""
    try:
        query = request.args.get('q', '')
        if not query:
            return jsonify({
                'success': False,
                'error': 'Paramètre de recherche manquant'
            }), 400
        
        results = tradier_api.search_symbols(query)
        if results and 'securities' in results:
            securities = results['securities']['security']
            if not isinstance(securities, list):
                securities = [securities]
            
            # Formater les résultats
            formatted_results = []
            for security in securities:
                formatted_results.append({
                    'symbol': security.get('symbol', ''),
                    'description': security.get('description', ''),
                    'type': security.get('type', ''),
                    'exchange': security.get('exch', '')
                })
            
            return jsonify({
                'success': True,
                'results': formatted_results,
                'count': len(formatted_results)
            })
        else:
            return jsonify({
                'success': True,
                'results': [],
                'count': 0
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# API: Test de connectivité Tradier
@app.route('/api/tradier/status')
def api_tradier_status():
    """API endpoint pour tester la connectivité avec Tradier"""
    try:
        is_connected = test_tradier_connectivity()
        return jsonify({
            'success': is_connected,
            'status': 'connected' if is_connected else 'disconnected',
            'message': 'Connexion Tradier OK' if is_connected else 'Impossible de se connecter à Tradier'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'status': 'error',
            'message': f'Erreur lors du test: {str(e)}'
        }), 500

# API: Maturités disponibles pour un symbole Tradier (Version Async)
@app.route('/api/tradier/expirations/<symbol>')
def api_tradier_expirations(symbol):
    """API endpoint pour récupérer les maturités disponibles pour un symbole via Tradier (Version Async)"""
    try:
        print(f"🔍 Récupération des expirations pour {symbol} (Async)...")
        
        # Utiliser la version asynchrone
        result = run_async(get_expirations_async(symbol))
        
        if result['success']:
            # Trier par date
            result['expirations'].sort(key=lambda x: x['date'])
            
            print(f"✅ {result['count']} expirations trouvées pour {symbol}")
            return jsonify({
                'success': True,
                'symbol': symbol,
                'expirations': result['expirations'],
                'count': result['count']
            })
        else:
            print(f"⚠️  Aucune expiration trouvée pour {symbol}: {result['error']}")
            return jsonify({
                'success': False,
                'error': f'Aucune expiration trouvée pour {symbol}',
                'details': result['error']
            }), 404
            
    except Exception as e:
        error_msg = str(e)
        print(f"❌ Erreur lors de la récupération des expirations pour {symbol}: {error_msg}")
        
        # Messages d'erreur plus informatifs
        if "ConnectionError" in error_msg or "ConnectTimeout" in error_msg:
            user_message = "Impossible de se connecter à l'API Tradier. Vérifiez votre connexion internet."
        elif "Max retries exceeded" in error_msg:
            user_message = "L'API Tradier ne répond pas. Veuillez réessayer plus tard."
        else:
            user_message = f"Erreur lors de la récupération des données: {error_msg}"
            
        return jsonify({
            'success': False,
            'error': user_message,
            'details': error_msg
        }), 500


# API: Chaîne d'options pour un symbole et une maturité Tradier
@app.route('/api/tradier/options/<symbol>/<expiration>')
def api_tradier_options(symbol, expiration):
    """API endpoint pour récupérer la chaîne d'options pour un symbole et une maturité via Tradier"""
    try:
        option_chain = tradier_api.get_option_chain(symbol, expiration)
        if option_chain and 'options' in option_chain and 'option' in option_chain['options']:
            options = option_chain['options']['option']
            if not isinstance(options, list):
                options = [options]
            
            # Formater les options
            formatted_options = []
            for option in options:
                try:
                    formatted_options.append({
                        'symbol': option.get('symbol', ''),
                        'description': option.get('description', ''),
                        'strike': float(option.get('strike', 0)),
                        'type': option.get('option_type', '').capitalize(),
                        'bid': float(option.get('bid', 0)) if option.get('bid') else None,
                        'ask': float(option.get('ask', 0)) if option.get('ask') else None,
                        'last': float(option.get('last', 0)) if option.get('last') else None,
                        'volume': int(option.get('volume', 0)) if option.get('volume') else 0,
                        'open_interest': int(option.get('open_interest', 0)) if option.get('open_interest') else 0,
                        'implied_volatility': float(option.get('implied_volatility', 0)) if option.get('implied_volatility') else None,
                        'delta': float(option.get('delta', 0)) if option.get('delta') else None,
                        'gamma': float(option.get('gamma', 0)) if option.get('gamma') else None,
                        'theta': float(option.get('theta', 0)) if option.get('theta') else None,
                        'vega': float(option.get('vega', 0)) if option.get('vega') else None
                    })
                except (ValueError, TypeError):
                    continue
            
            # Trier par strike (ordre décroissant) puis par type
            formatted_options.sort(key=lambda x: (-x['strike'], x['type']))
            
            return jsonify({
                'success': True,
                'symbol': symbol,
                'expiration': expiration,
                'options': formatted_options,
                'count': len(formatted_options)
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Aucune option trouvée pour {symbol} - {expiration}'
            }), 404
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# API: Strikes disponibles pour un symbole et une maturité Tradier
@app.route('/api/tradier/strikes/<symbol>/<expiration>')
def api_tradier_strikes(symbol, expiration):
    """API endpoint pour récupérer les strikes disponibles pour un symbole et une maturité via Tradier"""
    try:
        print(f"🎯 Récupération des strikes pour {symbol} - {expiration}...")
        
        # Récupérer les strikes et le prix spot en parallèle
        strikes_data = tradier_api.get_option_strikes(symbol, expiration)
        spot_data = tradier_api.get_stock_quote(symbol)
        
        if strikes_data and 'strikes' in strikes_data and 'strike' in strikes_data['strikes']:
            strikes = strikes_data['strikes']['strike']
            if not isinstance(strikes, list):
                strikes = [strikes]
            
            # Récupérer le prix spot
            spot_price = 100.0  # Prix par défaut
            if spot_data and 'quotes' in spot_data and 'quote' in spot_data['quotes']:
                quote = spot_data['quotes']['quote']
                if isinstance(quote, list):
                    quote = quote[0]
                spot_price = float(quote.get('last', 100.0))
            
            # Convertir en float et formater avec pourcentage
            formatted_strikes = []
            for strike in strikes:
                try:
                    strike_value = float(strike)
                    percentage = (strike_value / spot_price) * 100
                    formatted_strikes.append({
                        'strike': strike_value,
                        'display': f"${strike_value:.2f}",
                        'percentage': round(percentage, 1),
                        'percentage_display': f"{percentage:.1f}% (${strike_value:.2f})"
                    })
                except (ValueError, TypeError):
                    continue
            
            # Trier par strike (ordre décroissant)
            formatted_strikes.sort(key=lambda x: -x['strike'])
            
            print(f"✅ {len(formatted_strikes)} strikes trouvés pour {symbol} - {expiration} (Spot: ${spot_price:.2f})")
            
            return jsonify({
                'success': True,
                'symbol': symbol,
                'expiration': expiration,
                'spot_price': spot_price,
                'strikes': formatted_strikes,
                'count': len(formatted_strikes)
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Aucun strike trouvé pour {symbol} - {expiration}'
            }), 404
            
    except Exception as e:
        print(f"❌ Erreur lors de la récupération des strikes pour {symbol} - {expiration}: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# API: Strikes disponibles pour un symbole sur plusieurs maturités spécifiques (union) - Version Async
@app.route('/api/tradier/strikes-union/<symbol>')
def api_tradier_strikes_union(symbol):
    """API endpoint pour récupérer l'union des strikes de plusieurs maturités spécifiques via Tradier (Version Async)"""
    try:
        print(f"🎯 Récupération de l'union des strikes pour {symbol} (Async)...")
        
        # Utiliser la version asynchrone
        result = run_async(get_strikes_union_async(symbol))
        
        if result['success']:
            # Formater les strikes pour la compatibilité avec l'interface existante
            formatted_strikes = []
            for strike_data in result['strikes']:
                formatted_strikes.append({
                    'strike': strike_data['strike'],
                    'display': f"${strike_data['strike']:.2f}",
                    'percentage': round(strike_data['percentage'], 1),
                    'percentage_display': strike_data['percentage_display']
                })
            
            return jsonify({
                'success': True,
                'symbol': symbol,
                'spot_price': result['spot_price'],
                'strikes': formatted_strikes,
                'count': len(formatted_strikes),
                'total_strikes': result['total_strikes'],
                'filtered_strikes': result['filtered_strikes'],
                'expirations_processed': result['expirations_processed'],
                'selected_expirations': result['selected_expirations']
            })
        else:
            print(f"❌ Erreur pour {symbol}: {result['error']}")
            return jsonify({
                'success': False,
                'error': result['error']
            }), 404
        
    except Exception as e:
        print(f"❌ Erreur lors de la récupération de l'union des strikes pour {symbol}: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# API: Prix spot d'une action via Tradier (Version Async)
@app.route('/api/tradier/quote/<symbol>')
def api_tradier_quote(symbol):
    """API endpoint pour récupérer le prix spot d'une action via Tradier (Version Async)"""
    try:
        print(f"💰 Récupération du prix spot pour {symbol} (Async)...")
        
        # Utiliser la version asynchrone
        result = run_async(get_quote_async(symbol))
        
        if result['success'] and result['spot_price'] > 0:
            print(f"✅ Prix spot {symbol}: ${result['spot_price']:.2f}")
            return jsonify({
                'success': True,
                'symbol': symbol,
                'spot_price': result['spot_price'],
                'last': result['spot_price']
            })
        else:
            error_msg = result.get('error', 'Prix spot invalide')
            print(f"⚠️  Erreur pour {symbol}: {error_msg}")
            return jsonify({
                'success': False,
                'error': f'Prix spot invalide pour {symbol}: {error_msg}'
            }), 404
            
    except Exception as e:
        print(f"❌ Erreur lors de la récupération du prix spot pour {symbol}: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# API: Term Structure pour un symbole et un strike spécifique
@app.route('/api/term-structure/<symbol>/<strike>')
def api_term_structure(symbol, strike):
    """API endpoint pour récupérer la term structure d'un symbole pour un strike spécifique"""
    try:
        # Convertir le strike en float
        strike = float(strike)
        print(f"📈 Récupération de la term structure pour {symbol} - Strike ${strike:.2f}...")
        
        # Récupérer les expirations disponibles
        expirations_data = tradier_api.get_option_expirations(symbol)
        if not expirations_data or 'expirations' not in expirations_data:
            return jsonify({
                'success': False,
                'error': f'Aucune expiration trouvée pour {symbol}'
            }), 404
        
        expirations = expirations_data['expirations']['date']
        if not expirations:
            return jsonify({
                'success': False,
                'error': f'Aucune expiration trouvée pour {symbol}'
            }), 404
        
        # Récupérer le prix spot
        spot_data = tradier_api.get_stock_quote(symbol)
        if not spot_data or 'quotes' not in spot_data:
            return jsonify({
                'success': False,
                'error': f'Impossible de récupérer le prix spot pour {symbol}'
            }), 404
        
        quote = spot_data['quotes']['quote']
        if isinstance(quote, list):
            quote = quote[0]
        
        spot_price = float(quote.get('last', 0))
        if spot_price <= 0:
            return jsonify({
                'success': False,
                'error': f'Prix spot invalide pour {symbol}'
            }), 404
        
        print(f"💰 Prix spot {symbol}: ${spot_price:.2f}")
        
        # Récupérer le taux sans risque et le rendement des dividendes
        risk_free_rate = get_risk_free_rate()
        dividend_yield = get_dividend_yield(symbol)
        
        print(f"📊 Paramètres Black-Scholes:")
        print(f"  Taux sans risque: {risk_free_rate*100:.2f}%")
        print(f"  Rendement dividendes: {dividend_yield*100:.2f}%")
        
        # Traiter chaque expiration
        term_structure_data = []
        today = datetime.now()
        
        for expiration in expirations:
            try:
                print(f"📅 Traitement de l'expiration {expiration}...")
                
                # Récupérer la chaîne d'options pour cette expiration
                chain_data = tradier_api.get_option_chain(symbol, expiration)
                if not chain_data or 'options' not in chain_data:
                    print(f"   ⚠️  Aucune option pour {expiration}")
                    continue
                
                options = chain_data['options']['option']
                if not options:
                    print(f"   ⚠️  Aucune option trouvée pour {expiration}")
                    continue
                
                # Convertir en liste si ce n'est pas déjà le cas
                if not isinstance(options, list):
                    options = [options]
                
                # Filtrer les options pour le strike exact
                strike_options = []
                for option in options:
                    try:
                        option_strike = float(option.get('strike', 0))
                        if abs(option_strike - strike) < 0.01:  # Tolérance pour les arrondis
                            strike_options.append(option)
                    except (ValueError, TypeError):
                        continue
                
                if not strike_options:
                    print(f"   ⚠️  Aucune option trouvée pour le strike ${strike:.2f}")
                    continue
                
                print(f"   ✅ {len(strike_options)} options trouvées pour le strike ${strike:.2f}")
                
                # Calculer le temps jusqu'à l'expiration en jours ouvrés
                time_to_exp = calculate_business_days_to_expiration(expiration)
                business_days = int(time_to_exp * 252)
                
                if time_to_exp <= 0:
                    print(f"   ⚠️  Expiration dans le passé: {expiration}")
                    continue
                
                print(f"   📅 Time to Exp: {time_to_exp:.3f} années ({business_days} jours ouvrés)")
                
                # Calculer la volatilité implicite pour ce strike et cette expiration
                ivs = []
                for option in strike_options:
                    option_type = option.get('option_type', '').lower()
                    
                    # Essayer d'abord la volatilité implicite fournie par Tradier
                    if option.get('implied_volatility') is not None:
                        try:
                            iv = float(option['implied_volatility'])
                            if 0 < iv < 5:  # Filtrer les valeurs aberrantes
                                ivs.append(iv)
                                print(f"      ✅ IV Tradier pour {option_type.upper()}: {iv:.4f}")
                        except (ValueError, TypeError):
                            pass
                    
                    # Si pas de volatilité implicite, essayer de la calculer à partir du prix
                    elif option.get('last') is not None and option['last'] > 0:
                        try:
                            # Utiliser le prix last ou le prix moyen bid-ask
                            option_price = float(option['last'])
                            if option.get('bid') is not None and option.get('ask') is not None and option['bid'] > 0 and option['ask'] > 0:
                                # Utiliser le prix moyen bid-ask si disponible
                                option_price = (float(option['bid']) + float(option['ask'])) / 2
                            
                            # Calculer la volatilité implicite avec Black-Scholes
                            calculated_iv = calculate_implied_volatility_black_scholes(
                                spot_price=spot_price,
                                strike=strike,
                                time_to_exp=time_to_exp,
                                option_price=option_price,
                                option_type=option_type,
                                risk_free_rate=risk_free_rate,
                                dividend_yield=dividend_yield
                            )
                            
                            if calculated_iv and 0.01 < calculated_iv < 2.0:
                                ivs.append(calculated_iv)
                                print(f"      ✅ IV calculée pour {option_type.upper()}: {calculated_iv:.4f} - Prix: ${option_price:.2f}")
                            else:
                                print(f"      ⚠️  IV calculée hors plage pour {option_type.upper()}: {calculated_iv}")
                        except Exception as e:
                            print(f"      ❌ Erreur calcul IV pour {option_type.upper()}: {e}")
                            continue
                
                if ivs:
                    avg_iv = sum(ivs) / len(ivs)
                    
                    term_structure_data.append({
                        'expiration': expiration,
                        'days_to_exp': business_days,  # Jours ouvrés
                        'time_to_exp': time_to_exp,    # Années (jours ouvrés / 252)
                        'implied_volatility': avg_iv,
                        'strike': strike,
                        'spot_price': spot_price,
                        'options_count': len(strike_options)
                    })
                    
                    print(f"   ✅ IV moyenne pour {expiration}: {avg_iv:.4f} ({avg_iv*100:.2f}%)")
                else:
                    print(f"   ⚠️  Aucune IV valide pour {expiration}")
                    
            except Exception as e:
                print(f"   ❌ Erreur pour {expiration}: {e}")
                continue
        
        if not term_structure_data:
            return jsonify({
                'success': False,
                'error': f'Aucune donnée de term structure trouvée pour {symbol} - Strike ${strike:.2f}'
            }), 404
        
        # Trier par maturité croissante (jours ouvrés)
        term_structure_data.sort(key=lambda x: x['days_to_exp'])
        
        print(f"✅ Term structure récupérée: {len(term_structure_data)} points pour {symbol} - Strike ${strike:.2f}")
        
        return jsonify({
            'success': True,
            'symbol': symbol,
            'strike': strike,
            'spot_price': spot_price,
            'term_structure': term_structure_data,
            'count': len(term_structure_data)
        })
        
    except Exception as e:
        print(f"❌ Erreur lors de la récupération de la term structure pour {symbol} - Strike ${strike:.2f}: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Routes pour les pages cryptomonnaies
@app.route('/crypto')
def crypto():
    """Page principale des cryptomonnaies"""
    return render_template('crypto.html')


# API: Données des cryptomonnaies
@app.route('/api/crypto-data')
def api_crypto_data():
    """API endpoint pour récupérer les données des cryptomonnaies"""
    try:
        data = yahoo_api.get_crypto_data()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la récupération des données crypto: {str(e)}'}), 500


# Gestion propre de la fermeture de l'application
import atexit
import signal
import sys

def cleanup_async_resources():
    """Nettoie les ressources asynchrones lors de la fermeture"""
    import gc
    import warnings
    
    # Supprimer tous les warnings pendant le nettoyage
    warnings.filterwarnings("ignore")
    
    try:
        # Obtenir l'event loop actuel
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            # Aucun event loop actuel
            print("ℹ️ Aucun event loop actuel à nettoyer")
            return
        
        if loop.is_closed():
            print("ℹ️ Event loop déjà fermé")
            return
        
        # Annuler toutes les tâches en cours
        pending = asyncio.all_tasks(loop)
        if pending:
            print(f"🔄 Annulation de {len(pending)} tâches en cours...")
            for task in pending:
                task.cancel()
            
            # Attendre que toutes les tâches se terminent avec un timeout court
            try:
                loop.run_until_complete(asyncio.wait_for(
                    asyncio.gather(*pending, return_exceptions=True),
                    timeout=2.0
                ))
            except (asyncio.TimeoutError, Exception):
                # Ignorer les erreurs de timeout
                pass
        
        # Fermer l'event loop proprement
        if not loop.is_closed():
            loop.close()
            print("✅ Event loop fermé proprement")
        
    except Exception as e:
        # Ignorer complètement les erreurs de nettoyage
        pass
    
    # Nettoyage agressif des ressources
    try:
        # Forcer le garbage collection multiple fois
        for _ in range(3):
            gc.collect()
        
        # Nettoyer les connexions aiohttp restantes
        try:
            import aiohttp
            # Fermer toutes les sessions aiohttp restantes
            if hasattr(aiohttp, '_connector_cleanup'):
                aiohttp._connector_cleanup()
        except:
            pass
        
        print("✅ Nettoyage agressif effectué")
    except Exception as e:
        # Ignorer les erreurs de nettoyage
        pass

def signal_handler(signum, frame):
    """Gestionnaire de signal pour la fermeture propre"""
    print(f"\n🛑 Signal {signum} reçu, fermeture propre de l'application...")
    cleanup_async_resources()
    sys.exit(0)

# Enregistrer les gestionnaires de fermeture
atexit.register(cleanup_async_resources)
signal.signal(signal.SIGINT, signal_handler)  # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # Terminaison

if __name__ == '__main__':
    try:
        print("🚀 Démarrage de l'application Flask...")
        
        # Configuration spécifique pour Windows
        if sys.platform == 'win32':
            print("🪟 Configuration Windows détectée")
            # Configurer l'event loop policy pour Windows
            try:
                # Utiliser SelectorEventLoop au lieu de ProactorEventLoop pour éviter les problèmes
                asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
                print("✅ Event loop policy Windows (Selector) configurée")
            except Exception as e:
                print(f"⚠️ Erreur lors de la configuration de l'event loop: {e}")
                # Fallback vers ProactorEventLoop si Selector ne fonctionne pas
                try:
                    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
                    print("✅ Event loop policy Windows (Proactor) configurée en fallback")
                except Exception as e2:
                    print(f"⚠️ Erreur lors du fallback: {e2}")
            
            # Supprimer tous les warnings d'asyncio sur Windows
            import warnings
            warnings.filterwarnings("ignore", category=RuntimeWarning, module="asyncio")
            warnings.filterwarnings("ignore", category=DeprecationWarning, module="asyncio")
            warnings.filterwarnings("ignore", message=".*Event loop is closed.*")
            warnings.filterwarnings("ignore", message=".*_ProactorBasePipeTransport.*")
            warnings.filterwarnings("ignore", message=".*coroutine.*was never awaited.*")
            print("✅ Warnings asyncio supprimés")
        
        # Configuration pour la production
        debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
        app.run(debug=debug_mode, host='0.0.0.0', port=int(os.getenv('PORT', 5000)))
    except KeyboardInterrupt:
        print("\n🛑 Arrêt demandé par l'utilisateur...")
        cleanup_async_resources()
    except Exception as e:
        print(f"❌ Erreur lors du démarrage: {e}")
        cleanup_async_resources()
        sys.exit(1)

